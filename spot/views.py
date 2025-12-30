from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.db.models import Q, Sum, Count
from django.utils import timezone
from datetime import datetime
from django.template.loader import render_to_string
from django.conf import settings
from django.urls import reverse
import json
import logging
from decimal import Decimal
from datetime import timedelta, datetime, time
import os
import re
import calendar
import urllib.parse
from io import BytesIO
from PIL import Image
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST

# Optional dependency: openpyxl for Excel export
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    Workbook = None
    Font = PatternFill = Alignment = None
    DataValidation = None

# Optional dependency: reportlab for PDF export
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
except ImportError:
    SimpleDocTemplate = Paragraph = Table = TableStyle = Spacer = None
    getSampleStyleSheet = None
    A4 = None
    colors = None

from .models import (
    User, Campaign, Spot, SpotSchedule,
    TimeSlot, PricingRule, CampaignHistory, Notification,
    CorrespondenceThread, CorrespondenceMessage,
    AdvisorySession, ContactRequest, AdvisoryArticle, CaseStudy,
    CoverageRequest, CoverageAttachment, Journalist, Driver, CoverageAssignment, AssignmentLog,
    AssignmentNotificationCampaign
)
from .forms import (
    CustomUserCreationForm, CustomAuthenticationForm, CampaignForm,
    SpotForm, CostSimulatorForm, CampaignSpotForm,
    AdvisorWizardForm, ContactRequestForm, CoverageRequestForm
)


def home(request):
    """Page d'accueil"""
    # Si un diffuseur arrive sur la racine, redirigeons-le vers son interface dédiée
    if request.user.is_authenticated and hasattr(request.user, 'is_diffuser') and request.user.is_diffuser():
        return redirect('diffusion_home')
    context = {
        'total_campaigns': Campaign.objects.count(),
        'active_campaigns': Campaign.objects.filter(status='active').count(),
        'total_clients': User.objects.filter(role='client').count(),
        # Nouveaux indicateurs orientés "service commercial"
        'requests_in_progress_count': Campaign.objects.filter(
            status__in=['pending', 'approved', 'active']
        ).count(),
        'scheduled_broadcasts_count': SpotSchedule.objects.filter(
            is_broadcasted=False
        ).count(),
        'completed_services_count': SpotSchedule.objects.filter(
            is_broadcasted=True
        ).count(),
    }
    return render(request, 'spot/home.html', context)


def register(request):
    """Inscription des utilisateurs"""
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            messages.success(request, 'Votre compte a été créé avec succès !')
            return redirect('login')
    else:
        form = CustomUserCreationForm()
    
    return render(request, 'spot/register.html', {'form': form})


def user_login(request):
    """Connexion des utilisateurs"""
    if request.method == 'POST':
        form = CustomAuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                messages.success(request, f'Bienvenue, {user.username} !')
                next_url = request.GET.get('next')
                if next_url:
                    return redirect(next_url)
                if hasattr(user, 'is_diffuser') and user.is_diffuser():
                    return redirect('diffusion_home')
                if hasattr(user, 'is_editorial_manager') and user.is_editorial_manager():
                    return redirect('editorial_dashboard')
                return redirect('home')
    else:
        form = CustomAuthenticationForm()
    
    return render(request, 'spot/login.html', {'form': form})


def user_logout(request):
    """Déconnexion des utilisateurs"""
    logout(request)
    messages.info(request, 'Vous avez été déconnecté.')
    return redirect('home')


@login_required
def dashboard(request):
    """Tableau de bord principal"""
    user = request.user
    if hasattr(user, 'is_editorial_manager') and user.is_editorial_manager():
        return redirect('editorial_dashboard')
    if user.is_client():
        campaigns = Campaign.objects.filter(client=user).order_by('-created_at')[:5]
        recent_spots = Spot.objects.filter(campaign__client=user).order_by('-created_at')[:5]
        notifications = Notification.objects.filter(
            user=user, 
            is_read=False
        ).order_by('-created_at')[:10]
        spots_in_progress_count = SpotSchedule.objects.filter(
            spot__campaign__client=user,
            is_broadcasted=False
        ).count()
        broadcasted_spots_count = SpotSchedule.objects.filter(
            spot__campaign__client=user,
            is_broadcasted=True
        ).count()
        context = {
            'campaigns': campaigns,
            'recent_spots': recent_spots,
            'notifications': notifications,
            'total_campaigns': Campaign.objects.filter(client=user).count(),
            'active_campaigns': Campaign.objects.filter(client=user, status='active').count(),
            'spots_in_progress_count': spots_in_progress_count,
            'broadcasted_spots_count': broadcasted_spots_count,
        }
    else:
        pending_campaigns = Campaign.objects.filter(status='pending').order_by('-created_at')
        pending_spots = Spot.objects.filter(status='pending_review').order_by('-created_at')
        notifications = Notification.objects.filter(
            user=user, 
            is_read=False
        ).order_by('-created_at')[:10]
        stats = {
            'total_campaigns': Campaign.objects.count(),
            'pending_campaigns': pending_campaigns.count(),
            'total_clients': User.objects.filter(role='client').count(),
        }
        context = {
            'pending_campaigns': pending_campaigns,
            'pending_spots': pending_spots,
            'notifications': notifications,
            'stats': stats,
        }
    return render(request, 'spot/dashboard.html', context)


@login_required
def campaign_list(request):
    """Liste des campagnes"""
    user = request.user
    
    if user.is_client():
        campaigns = Campaign.objects.filter(client=user)
    else:
        campaigns = Campaign.objects.all()
    
    # Filtres
    status_filter = request.GET.get('status')
    if status_filter:
        campaigns = campaigns.filter(status=status_filter)
    
    search = request.GET.get('search')
    if search:
        campaigns = campaigns.filter(
            Q(title__icontains=search) | 
            Q(description__icontains=search)
        )
    
    # Pagination
    paginator = Paginator(campaigns.order_by('-created_at'), 10)
    page_number = request.GET.get('page')
    campaigns = paginator.get_page(page_number)
    
    context = {
        'campaigns': campaigns,
        'status_choices': Campaign.STATUS_CHOICES,
    }
    
    return render(request, 'spot/campaign_list.html', context)


@login_required
def campaign_create(request):
    """Création d'une nouvelle campagne"""
    # Blocage préventif pour administrateurs
    if request.user.is_authenticated and hasattr(request.user, 'is_admin') and request.user.is_admin():
        logging.getLogger('bf1tv').info(
            'ADMIN_BLOCK_VIEW | user=%s username=%s view=campaign_create method=%s at=%s',
            getattr(request.user, 'id', None), getattr(request.user, 'username', ''), request.method, timezone.now().isoformat(timespec='seconds')
        )
        messages.error(request, "Cette fonctionnalité est réservée aux clients. En tant qu'administrateur, vous n'avez pas accès à cette option.")
        return redirect('dashboard')
    # Variables de contexte par défaut
    locked_fields = []
    selection_note = ''

    if request.method == 'POST':
        form = CampaignForm(request.POST)
        if form.is_valid():
            campaign = form.save(commit=False)
            campaign.client = request.user
            campaign.status = 'pending'
            campaign.save()
            form.save_m2m()  # Persiste preferred_time_slots
            if campaign.campaign_type == 'spot_upload':
                messages.success(request, 'Campagne créée avec succès ! Veuillez maintenant télécharger votre spot.')
                return redirect('spot_upload', campaign_id=campaign.id)
            else:
                messages.success(request, 'Campagne créée avec succès ! Notre équipe vous contactera pour la création du spot.')
                return redirect('campaign_detail', campaign_id=campaign.id)
        else:
            messages.error(request, 'Veuillez corriger les erreurs du formulaire.')
    else:
        # Pré-remplissage depuis la page Tarifs
        initial = {}
        locked_fields = []
        selection_note = request.GET.get('selection_note') or ''
        lock = (request.GET.get('lock') or '').strip()
        if lock:
            locked_fields = [f.strip() for f in lock.split(',') if f.strip()]

        # Champs simples (pré-remplis)
        for key in ['channel', 'campaign_type', 'objective', 'budget', 'title']:
            val = request.GET.get(key)
            if val:
                initial[key] = val

        # Description indicative depuis la sélection
        if selection_note and not request.GET.get('description'):
            initial['description'] = selection_note

        form = CampaignForm(initial=initial)

    return render(request, 'spot/campaign_create.html', {
        'form': form,
        'locked_fields': locked_fields,
        'selection_note': selection_note,
    })


@login_required
def campaign_spot_create(request):
    """Mode Facile: création de campagne + upload du spot en une fois"""
    # Blocage préventif pour administrateurs
    if request.user.is_authenticated and hasattr(request.user, 'is_admin') and request.user.is_admin():
        logging.getLogger('bf1tv').info(
            'ADMIN_BLOCK_VIEW | user=%s username=%s view=campaign_spot_create method=%s at=%s',
            getattr(request.user, 'id', None), getattr(request.user, 'username', ''), request.method, timezone.now().isoformat(timespec='seconds')
        )
        messages.error(request, "Cette fonctionnalité est réservée aux clients. En tant qu'administrateur, vous n'avez pas accès à cette option.")
        return redirect('dashboard')
    # Variables de contexte par défaut
    locked_fields = []
    selection_note = ''
    slot_label = ''

    if request.method == 'POST':
        form = CampaignSpotForm(request.POST, request.FILES)
        if form.is_valid():
            campaign = Campaign(
                title=form.cleaned_data['title'],
                description=form.cleaned_data['description'],
                start_date=form.cleaned_data['start_date'],
                end_date=form.cleaned_data['end_date'],
                budget=form.cleaned_data['budget'],
                client=request.user,
                status='pending'
             )
            campaign.objective = form.cleaned_data['objective']
            campaign.channel = form.cleaned_data['channel']
            campaign.languages = (form.cleaned_data.get('languages') or '').strip()
            campaign.campaign_type = form.cleaned_data.get('campaign_type') or 'spot_upload'
            campaign.requested_creation = bool(form.cleaned_data.get('requested_creation'))
            campaign.target_audience = form.cleaned_data.get('target_audience') or ''
            campaign.key_message = form.cleaned_data.get('key_message') or ''
            campaign.save()
            preferred = form.cleaned_data.get('preferred_time_slots')
            if preferred:
                campaign.preferred_time_slots.set(preferred)

            CampaignHistory.objects.create(
                campaign=campaign,
                action='created',
                description='Campagne créée via Mode Facile',
                user=request.user
            )

            # Spot: uniquement si l’utilisateur a fourni un titre ou un fichier
            spot_title = form.cleaned_data.get('spot_title')
            spot_desc = form.cleaned_data.get('spot_description', '')
            media_type = form.cleaned_data.get('media_type')
            image_file = form.cleaned_data.get('image_file')
            video_file = form.cleaned_data.get('video_file')
            duration_seconds = form.cleaned_data.get('duration_seconds')

            if spot_title or image_file or video_file:
                spot = Spot(
                    campaign=campaign,
                    title=spot_title,
                    description=spot_desc,
                    status='pending_review'
                )
                if video_file:
                    spot.media_type = 'video'
                    spot.video_file = video_file
                    spot.duration_seconds = duration_seconds
                elif image_file:
                    spot.media_type = 'image'
                    spot.image_file = image_file
                else:
                    spot.media_type = media_type or 'image'
                spot.save()
                messages.success(request, "Votre pub est enregistrée. Elle sera validée par un administrateur.")
            else:
                messages.success(request, "Votre campagne est enregistrée. Vous pourrez uploader votre spot plus tard.")
            return redirect('campaign_detail', campaign_id=campaign.id)
        else:
            messages.error(request, "Veuillez corriger les erreurs du formulaire.")
    else:
        today = timezone.now().date()
        initial = {
            'start_date': today,
            'end_date': today + timedelta(days=14),
        }
        # Pré-remplissage depuis la page Tarifs
        locked_fields = []
        selection_note = request.GET.get('selection_note') or ''
        lock = (request.GET.get('lock') or '').strip()
        if lock:
            locked_fields = [f.strip() for f in lock.split(',') if f.strip()]
        for key in ['channel', 'campaign_type', 'objective', 'duration_seconds', 'budget', 'title', 'spot_title', 'media_type']:
            val = request.GET.get(key)
            if val:
                initial[key] = val
        # Tentative: mapper slot_label vers un TimeSlot existant pour M2M
        slot_label = (request.GET.get('slot_label') or '').strip()
        preferred_ids = []
        if slot_label:
            try:
                # Ex: "6h – 11h"
                m = re.search(r'(\d{1,2})h.*?(\d{1,2})h', slot_label)
                if m:
                    s = time(int(m.group(1)), 0)
                    e = time(int(m.group(2)), 0)
                    ts = TimeSlot.objects.filter(start_time=s, end_time=e).first()
                    if ts:
                        preferred_ids = [ts.id]
            except Exception:
                pass
        if preferred_ids:
            initial['preferred_time_slots'] = preferred_ids
        # Décrire par défaut avec la sélection
        if selection_note and not request.GET.get('description'):
            initial['description'] = selection_note
            initial['spot_description'] = selection_note
        form = CampaignSpotForm(initial=initial)

    return render(request, 'spot/campaign_spot_create.html', {
        'form': form,
        'title': 'Campagne/Spot',
        'locked_fields': locked_fields,
        'selection_note': selection_note,
        'pref_slot_label': slot_label if 'slot_label' in request.GET else '',
    })


@login_required
def coverage_request_create(request):
    """Demande de couverture médiatique (wizard multi-étapes)"""
    # Bloquer les administrateurs (réservé aux clients)
    if request.user.is_authenticated and hasattr(request.user, 'is_admin') and request.user.is_admin():
        logging.getLogger('bf1tv').info(
            'ADMIN_BLOCK_VIEW | user=%s username=%s view=coverage_request_create method=%s at=%s',
            getattr(request.user, 'id', None), getattr(request.user, 'username', ''), request.method, timezone.now().isoformat(timespec='seconds')
        )
        messages.error(request, "Cette fonctionnalité est réservée aux clients. En tant qu'administrateur, vous n'avez pas accès à cette option.")
        return redirect('dashboard')

    if request.method == 'POST':
        form = CoverageRequestForm(request.POST, request.FILES)
        if form.is_valid():
            coverage = form.save(commit=False)
            coverage.user = request.user
            coverage.status = 'new'
            coverage.save()
            # Pièces jointes multiples
            for f in request.FILES.getlist('attachments'):
                try:
                    CoverageAttachment.objects.create(request=coverage, file=f)
                except Exception:
                    # Continuer même si une pièce jointe pose problème
                    pass
            # Notifications aux administrateurs
            admins = User.objects.filter(role='admin')
            for admin in admins:
                Notification.objects.create(
                    user=admin,
                    title='Nouvelle demande de couverture',
                    message=f"{request.user.username} a soumis une demande: '{coverage.event_title}'",
                    type='info'
                )
            messages.success(request, 'Votre demande de couverture a été envoyée. Notre rédaction vous contactera rapidement.')
            return redirect('dashboard')
        else:
            messages.error(request, 'Veuillez corriger les erreurs du formulaire.')
    else:
        # Pré-remplissage depuis paramètres GET (Tarifs / redirections internes)
        initial = {}
        selection_note = request.GET.get('selection_note') or ''

        # Champs principaux
        initial['event_title'] = request.GET.get('event_title') or request.GET.get('title') or ''
        initial['description'] = request.GET.get('description') or selection_note or ''
        initial['coverage_type'] = request.GET.get('coverage_type') or 'editorial_defined'
        initial['coverage_objective'] = request.GET.get('coverage_objective') or selection_note or ''
        initial['urgency_level'] = request.GET.get('urgency_level') or 'normal'
        initial['event_type'] = request.GET.get('event_type') or 'other'

        # Coordonnées et logistique (si présentes)
        for key in [
            'contact_name','contact_phone','contact_email','other_contacts',
            'address','meeting_point','response_deadline',
            'event_date','start_time','end_time'
        ]:
            val = request.GET.get(key)
            if val:
                initial[key] = val

        form = CoverageRequestForm(initial=initial)

    return render(request, 'spot/coverage_request_create.html', {
        'form': form,
        'title': 'Demande de couverture médiatique',
    })


@login_required
def campaign_detail(request, campaign_id):
    """Détails d'une campagne"""
    campaign = get_object_or_404(Campaign, id=campaign_id)
    
    # Vérifier les permissions
    if request.user.is_client() and campaign.client != request.user:
        messages.error(request, 'Vous n\'avez pas accès à cette campagne.')
        return redirect('campaign_list')
    
    spots = campaign.spots.all()
    # Filtrer l'historique pour ne garder que les événements utiles
    history_qs = campaign.history.filter(
        action__in=['approved', 'rejected', 'spot_uploaded', 'spot_approved', 'broadcasted']
    ).order_by('-created_at')
    history = list(history_qs)
    
    context = {
        'campaign': campaign,
        'spots': spots,
        'history': history,
        'history_display_limit': 5,
        'suppress_messages': True,
    }
    
    return render(request, 'spot/campaign_detail.html', context)


@login_required
def spot_upload(request, campaign_id):
    """Téléchargement d'un spot"""
    campaign = get_object_or_404(Campaign, id=campaign_id)
    
    # Vérifier les permissions
    if request.user.is_client() and campaign.client != request.user:
        messages.error(request, 'Vous n\'avez pas accès à cette campagne.')
        return redirect('campaign_list')
    
    # Vérifier si la campagne a déjà un spot
    existing_spot = campaign.spots.first()
    if existing_spot:
        messages.info(request, 'Cette campagne a déjà un spot.')
        return redirect('campaign_detail', campaign_id=campaign.id)
    
    if request.method == 'POST':
        form = SpotForm(request.POST, request.FILES)
        if form.is_valid():
            spot = form.save(commit=False)
            spot.campaign = campaign
            spot.status = 'pending_review'
            spot.save()
            
            # Mettre à jour le statut de la campagne si nécessaire
            if campaign.status == 'draft':
                campaign.status = 'pending'
                campaign.save()
            
            # Créer une notification pour les administrateurs
            admins = User.objects.filter(role='admin')
            for admin in admins:
                Notification.objects.create(
                    user=admin,
                    title='Nouveau spot à approuver',
                    message=f'Un nouveau spot a été téléchargé pour la campagne "{campaign.title}".',
                    related_campaign=campaign,
                    related_spot=spot
                )
            
            messages.success(request, 'Votre spot a été téléchargé avec succès et sera examiné par notre équipe.')
            return redirect('campaign_detail', campaign_id=campaign.id)
    else:
        form = SpotForm()
    
    context = {
        'form': form,
        'campaign': campaign,
    }
    
    return render(request, 'spot/spot_upload.html', context)


@login_required
def notifications(request):
    """Liste des notifications"""
    notifications = Notification.objects.filter(user=request.user).order_by('-created_at')
    unread_count = Notification.objects.filter(user=request.user, is_read=False).count()
    
    # Pagination
    paginator = Paginator(notifications, 20)
    page_number = request.GET.get('page')
    notifications = paginator.get_page(page_number)
    
    return render(request, 'spot/notifications.html', {'notifications': notifications, 'unread_count': unread_count})


@login_required
def notifications_mark_read(request, id):
    """Marque une notification comme lue et retourne le compteur mis à jour."""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Méthode invalide'}, status=405)
    notification = get_object_or_404(Notification, id=id, user=request.user)
    if not notification.is_read:
        notification.is_read = True
        notification.save(update_fields=['is_read'])
    unread_count = Notification.objects.filter(user=request.user, is_read=False).count()
    return JsonResponse({'status': 'success', 'unread_count': unread_count})


@login_required
def notifications_mark_all_read(request):
    """Marque toutes les notifications de l'utilisateur comme lues."""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Méthode invalide'}, status=405)
    Notification.objects.filter(user=request.user, is_read=False).update(is_read=True)
    return JsonResponse({'status': 'success', 'unread_count': 0})


@login_required
def notifications_list_partial(request):
    """Retourne la liste HTML des notifications (partielle) et le compteur non lu."""
    notifications = Notification.objects.filter(user=request.user).order_by('-created_at')
    paginator = Paginator(notifications, 20)
    page_number = request.GET.get('page')
    notifications = paginator.get_page(page_number)
    html = render_to_string('spot/includes/notifications_list.html', {'notifications': notifications}, request=request)
    unread_count = Notification.objects.filter(user=request.user, is_read=False).count()
    return JsonResponse({'status': 'success', 'html': html, 'unread_count': unread_count})


@login_required
def notifications_delete(request, id):
    """Supprime une notification de l'utilisateur et retourne le compteur mis à jour."""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Méthode invalide'}, status=405)
    notification = get_object_or_404(Notification, id=id, user=request.user)
    notification.delete()
    unread_count = Notification.objects.filter(user=request.user, is_read=False).count()
    return JsonResponse({'status': 'success', 'unread_count': unread_count})


# Vues pour les administrateurs
from functools import wraps

def admin_required(view_func):
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('admin_login')
        if not request.user.is_admin():
            messages.error(request, 'Accès réservé aux administrateurs.')
            return redirect('dashboard')
        return view_func(request, *args, **kwargs)
    return _wrapped

@admin_required
def admin_contact_request_detail(request, request_id):
    """Détail d'une demande de contact et action de réponse admin."""
    contact = get_object_or_404(ContactRequest, id=request_id)
    thread = None

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'respond':
            # Marquer la demande comme 'contacted' et l'assigner
            contact.status = 'contacted'
            contact.assigned_to = request.user
            contact.save(update_fields=['status', 'assigned_to', 'updated_at'])
            
            # Si l'utilisateur existe, créer/ouvrir un fil et envoyer un message
            if contact.user:
                thread = CorrespondenceThread.objects.create(
                    client=contact.user,
                    subject=contact.subject,
                    related_campaign=None,
                    status='open',
                    last_message_at=timezone.now()
                )
                CorrespondenceMessage.objects.create(
                    thread=thread,
                    author=request.user,
                    content="Bonjour, votre demande a bien été reçue. Un conseiller vous recontactera très bientôt."
                )
                Notification.objects.create(
                    user=contact.user,
                    title='Réponse du support',
                    message=f'Votre demande "{contact.subject}" a été prise en compte. Un conseiller vous recontactera.',
                    related_thread=thread
                )
                messages.success(request, "Réponse envoyée et fil de discussion créé.")
                return redirect('correspondence_thread', thread_id=thread.id)
            else:
                messages.success(request, "Demande marquée comme contactée. (Utilisateur non connecté)")
                return redirect('admin_dashboard')

    return render(request, 'spot/admin_contact_request_detail.html', {
        'contact': contact,
        'thread': thread
    })

def admin_campaign_approve(request, campaign_id):
    """Approbation d'une campagne par l'admin"""
    campaign = get_object_or_404(Campaign, id=campaign_id)
    # Retirer la contrainte 'spot obligatoire' et gérer si spot présent ou non
    spot = campaign.spots.first()
    
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'approve':
            # Approuver la campagne seule; approuver le spot uniquement s'il existe
            campaign.status = 'approved'
            campaign.approved_by = request.user
            campaign.approved_at = timezone.now()
            campaign.save()
            
            if spot:
                spot.status = 'approved'
                spot.approved_by = request.user
                spot.approved_at = timezone.now()
                spot.save()
            
            messages.success(request, 'Campagne approuvée' + (' et spot approuvé !' if spot else ' !'))
        elif action == 'reject':
            rejection_reason = request.POST.get('rejection_reason', '')
            
            campaign.status = 'rejected'
            campaign.approved_by = request.user
            campaign.approved_at = timezone.now()
            campaign.rejection_reason = rejection_reason
            campaign.save()
            
            if spot:
                spot.status = 'rejected'
                spot.approved_by = request.user
                spot.approved_at = timezone.now()
                spot.rejection_reason = rejection_reason
                spot.save()
            
            messages.success(request, 'Campagne rejetée' + (' et spot rejeté !' if spot else ' !'))
        
        return redirect('admin_dashboard')
    
    context = {
        'campaign': campaign,
        'spot': spot
    }
    
    return render(request, 'spot/admin_campaign_approve.html', context)


@login_required
def admin_spot_approve(request, spot_id):
    """Approbation d'un spot par l'admin"""
    spot = get_object_or_404(Spot, id=spot_id)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'approve':
            spot.status = 'approved'
            spot.approved_by = request.user
            spot.approved_at = timezone.now()
            spot.save()
            messages.success(request, 'Spot approuvé !')
        elif action == 'reject':
            spot.status = 'rejected'
            spot.approved_by = request.user
            spot.approved_at = timezone.now()
            spot.rejection_reason = request.POST.get('rejection_reason', '')
            spot.save()
            messages.success(request, 'Spot rejeté !')
        
        return redirect('admin_dashboard')
    
    return render(request, 'spot/admin_spot_approve.html', {'spot': spot})


@login_required
def admin_campaign_list(request):
    status = request.GET.get('status')
    q = request.GET.get('q', '').strip()
    campaigns = Campaign.objects.all().order_by('-created_at')
    if status:
        campaigns = campaigns.filter(status=status)
    if q:
        campaigns = campaigns.filter(Q(title__icontains=q) | Q(client__username__icontains=q))
    paginator = Paginator(campaigns, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    status_counts = Campaign.objects.values('status').annotate(count=Count('status'))
    return render(request, 'spot/admin_campaign_list.html', {
        'page_obj': page_obj,
        'status': status,
        'q': q,
        'status_counts': {item['status']: item['count'] for item in status_counts},
    })


@admin_required
def admin_coverage_list(request):
    """Liste des demandes de couverture (admin) avec actions de statut"""
    status = request.GET.get('status')
    q = (request.GET.get('q') or '').strip()

    qs = CoverageRequest.objects.select_related('user').order_by('-created_at')
    if status:
        qs = qs.filter(status=status)
    if q:
        qs = qs.filter(
            Q(event_title__icontains=q) |
            Q(contact_name__icontains=q) |
            Q(contact_email__icontains=q)
        )

    if request.method == 'POST':
        action = request.POST.get('action')
        coverage_id = request.POST.get('coverage_id')
        cv = CoverageRequest.objects.filter(id=coverage_id).first()
        if cv:
            if action == 'review':
                cv.status = 'review'
            elif action == 'schedule':
                cv.status = 'scheduled'
            elif action == 'close':
                cv.status = 'closed'
            cv.save(update_fields=['status', 'updated_at'])
            messages.success(request, 'Statut mis à jour.')
        return redirect('admin_coverage_list')

    paginator = Paginator(qs, 20)
    page_obj = paginator.get_page(request.GET.get('page'))
    status_counts = CoverageRequest.objects.values('status').annotate(count=Count('status'))

    return render(request, 'spot/admin_coverage_list.html', {
        'page_obj': page_obj,
        'status': status,
        'q': q,
        'status_counts': {item['status']: item['count'] for item in status_counts},
    })


@admin_required
def admin_coverage_detail(request, coverage_id):
    """Détail d'une demande de couverture pour administrateur.
    Affiche toutes les informations et permet d'approuver ou de rejeter
    avec commentaire. Les décisions sont auditées via notifications et logs.
    """
    coverage = get_object_or_404(CoverageRequest.objects.select_related('user'), id=coverage_id)
    attachments = coverage.attachments.all()

    if request.method == 'POST':
        action = request.POST.get('action')
        comment = (request.POST.get('comment') or '').strip()

        if action == 'review':
            coverage.status = 'review'
            coverage.save(update_fields=['status', 'updated_at'])
            messages.success(request, 'La demande est passée en révision.')
            # Audit via notification pour l'admin et le client
            Notification.objects.create(
                user=request.user,
                title='Demande de couverture en révision',
                message=f"{coverage.event_title} est passée en révision.",
                type='info'
            )
            if coverage.user:
                Notification.objects.create(
                    user=coverage.user,
                    title='Votre demande est en cours de révision',
                    message=f"Votre demande pour '{coverage.event_title}' est en cours de révision.",
                    type='info'
                )
        elif action == 'approve':
            coverage.status = 'scheduled'
            coverage.save(update_fields=['status', 'updated_at'])
            messages.success(request, 'La demande a été validée et planifiée.')
            if coverage.user:
                Notification.objects.create(
                    user=coverage.user,
                    title='Demande de couverture validée',
                    message=f"Votre demande pour '{coverage.event_title}' a été validée.",
                    type='success',
                    related_coverage=coverage
                )
        elif action == 'reject':
            if not comment:
                messages.error(request, 'La raison du rejet est obligatoire.')
                return render(request, 'spot/admin_coverage_detail.html', {
                    'coverage': coverage,
                    'attachments': attachments,
                    'title': f"Demande de couverture — {coverage.event_title}",
                    'show_reject': True,
                })
            coverage.status = 'closed'
            coverage.save(update_fields=['status', 'updated_at'])
            messages.warning(request, 'La demande a été rejetée.')
            if coverage.user:
                Notification.objects.create(
                    user=coverage.user,
                    title='Demande de couverture rejetée',
                    message=f"Votre demande pour '{coverage.event_title}' a été rejetée. Motif: {comment}",
                    type='error',
                    related_coverage=coverage
                )
        else:
            messages.error(request, 'Action inconnue.')

        # Journaliser la décision dans les logs applicatifs
        try:
            logger = logging.getLogger('bf1tv') if 'bf1tv' in logging.Logger.manager.loggerDict else logging.getLogger(__name__)
            logger.info(
                f"CoverageRequest {coverage.id} action={action} by={request.user.username} comment='{comment}'"
            )
        except Exception:
            pass

        return redirect('admin_coverage_detail', coverage_id=coverage.id)

    return render(request, 'spot/admin_coverage_detail.html', {
        'coverage': coverage,
        'attachments': attachments,
        'title': f"Demande de couverture — {coverage.event_title}",
    })


@login_required
def coverage_request_detail(request, coverage_id):
    """Détail d'une demande de couverture pour le client.
    Accessible au propriétaire de la demande (coverage.user) ou à un admin.
    """
    coverage = get_object_or_404(CoverageRequest.objects.select_related('user'), id=coverage_id)
    # Contrôle d'accès
    if not (request.user.is_admin() or (coverage.user and coverage.user == request.user)):
        messages.error(request, "Vous n'avez pas accès à cette demande de couverture.")
        return redirect('dashboard')
    attachments = coverage.attachments.all()
    return render(request, 'spot/coverage_request_detail.html', {
        'coverage': coverage,
        'attachments': attachments,
        'title': f"Ma demande de couverture — {coverage.event_title}",
    })


@login_required
def admin_dashboard(request):
    """Tableau de bord administrateur"""
    if not request.user.is_admin():
        messages.error(request, 'Accès non autorisé.')
        return redirect('dashboard')
    stats = {
        'total_campaigns': Campaign.objects.count(),
        'pending_campaigns': Campaign.objects.filter(status='pending').count(),
        'active_campaigns': Campaign.objects.filter(status='active').count(),
        'total_clients': User.objects.filter(role='client').count(),
        # Compteurs pour bulles d'indicateurs
        'pending_spots': Spot.objects.filter(status='pending_review').count(),
        'pending_threads': CorrespondenceThread.objects.filter(status='pending').count(),
        'pending_coverages': CoverageRequest.objects.filter(status='new').count(),
    }
    recent_campaigns = Campaign.objects.order_by('-created_at')[:10]
    pending_approvals = {
        'campaigns': Campaign.objects.filter(status='pending'),
        'spots': Spot.objects.filter(status='pending_review'),
        'coverages': CoverageRequest.objects.filter(status='new'),
    }
    # Données complémentaires pour sections Alertes & Activités
    notifications = Notification.objects.filter(user=request.user).order_by('-created_at')[:10]
    recent_actions = CampaignHistory.objects.order_by('-created_at')[:20]
    context = {
        'stats': stats,
        'recent_campaigns': recent_campaigns,
        'pending_approvals': pending_approvals,
        'notifications': notifications,
        'recent_actions': recent_actions,
    }
    return render(request, 'spot/admin_dashboard.html', context)


@login_required
def spot_detail(request, spot_id):
    spot = get_object_or_404(Spot, id=spot_id)
    
    # Vérifier que l'utilisateur est propriétaire ou admin
    if not (request.user == spot.campaign.client or request.user.is_admin()):
        messages.error(request, "Vous n'avez pas accès à ce spot.")
        return redirect('dashboard')
    
    context = {
        'spot': spot,
        'campaign': spot.campaign,
        'title': f'Spot: {spot.title}'
    }
    
    return render(request, 'spot/spot_detail.html', context)

@login_required
def admin_logout(request):
    """Déconnexion des administrateurs"""
    logout(request)
    messages.info(request, "Vous avez été déconnecté de l'administration.")
    return redirect('admin_login')

@login_required
def spot_list(request):
    """Liste des spots (client: ses spots; admin: tous)"""
    if request.user.is_admin():
        qs = Spot.objects.select_related('campaign').order_by('-created_at')
    else:
        qs = Spot.objects.filter(campaign__client=request.user).select_related('campaign').order_by('-created_at')

    status = request.GET.get('status')
    valid_statuses = dict(Spot.STATUS_CHOICES)
    if status in valid_statuses:
        qs = qs.filter(status=status)

    paginator = Paginator(qs, 10)
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(request, 'spot/spot_list.html', {
        'spots': page_obj.object_list,
        'page_obj': page_obj,
        'status': status,
        'status_choices': Spot.STATUS_CHOICES,
        'title': 'Mes Spots' if request.user.is_client() else 'Tous les spots'
    })

@login_required
def broadcast_grid(request):
    """Grille de diffusion par date, regroupée par créneau (TimeSlot)"""
    selected_date_str = request.GET.get('date')
    try:
        selected_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date() if selected_date_str else timezone.now().date()
    except Exception:
        selected_date = timezone.now().date()

    # Créneaux actifs
    time_slots = TimeSlot.objects.filter(is_active=True).order_by('start_time')

    # Filtre créneau optionnel
    slot_id = request.GET.get('slot_id')
    if slot_id:
        time_slots = time_slots.filter(id=slot_id)

    # Programmes de la date choisie
    qs = SpotSchedule.objects.select_related('spot', 'time_slot', 'spot__campaign')\
        .filter(broadcast_date=selected_date)
    if request.user.is_client():
        qs = qs.filter(spot__campaign__client=request.user)
    qs = qs.order_by('broadcast_time')

    # Regroupement par créneau
    grid = []
    total_schedules = 0
    for ts in time_slots:
        entries = list(qs.filter(time_slot=ts))
        total_schedules += len(entries)
        grid.append({
            'slot': ts,
            'entries': entries,
            'count': len(entries),
        })

    # Campagnes couvrant la date (alignement avec créations)
    campaigns_qs = Campaign.objects.filter(start_date__lte=selected_date, end_date__gte=selected_date)
    if request.user.is_client():
        campaigns_qs = campaigns_qs.filter(client=request.user)
    campaigns_qs = campaigns_qs.select_related('client').order_by('-created_at')

    # Calculs par campagne: spots et programmations du jour
    campaigns_for_day = []
    for c in campaigns_qs:
        spots_count = c.spots.count()
        schedules_count = SpotSchedule.objects.filter(spot__campaign=c, broadcast_date=selected_date).count()
        campaigns_for_day.append({
            'campaign': c,
            'spots_count': spots_count,
            'schedules_count': schedules_count,
        })

    # Navigation date précédente/suivante
    prev_date = (selected_date - timedelta(days=1)).strftime('%Y-%m-%d')
    next_date = (selected_date + timedelta(days=1)).strftime('%Y-%m-%d')

    context = {
        'selected_date': selected_date,
        'time_slots': time_slots,
        'grid': grid,
        'total_schedules': total_schedules,
        'slot_id': slot_id,
        'prev_date': prev_date,
        'next_date': next_date,
        'user_is_client': request.user.is_client(),
        'campaigns_for_day': campaigns_for_day,
        'campaigns_for_day_count': len(campaigns_for_day),
    }
    return render(request, 'spot/broadcast_grid.html', context)

@login_required
def correspondence_list(request):
    """Liste des correspondances (tickets)"""
    if request.user.is_admin():
        threads = CorrespondenceThread.objects.select_related('client', 'related_campaign')\
            .prefetch_related('messages')\
            .order_by('-last_message_at', '-created_at')
    else:
        threads = CorrespondenceThread.objects.filter(client=request.user)\
            .select_related('related_campaign')\
            .prefetch_related('messages')\
            .order_by('-last_message_at', '-created_at')

    # Filtre par statut
    status = request.GET.get('status')
    valid_statuses = dict(CorrespondenceThread.STATUS_CHOICES)
    if status in valid_statuses:
        threads = threads.filter(status=status)

    # Recherche par sujet
    q = (request.GET.get('q') or '').strip()
    if q:
        threads = threads.filter(subject__icontains=q)

    # Compteurs
    if request.user.is_admin():
        stats = {
            'open': CorrespondenceThread.objects.filter(status='open').count(),
            'pending': CorrespondenceThread.objects.filter(status='pending').count(),
            'closed': CorrespondenceThread.objects.filter(status='closed').count(),
        }
    else:
        stats = {
            'open': CorrespondenceThread.objects.filter(client=request.user, status='open').count(),
            'pending': CorrespondenceThread.objects.filter(client=request.user, status='pending').count(),
            'closed': CorrespondenceThread.objects.filter(client=request.user, status='closed').count(),
        }

    # Nouveau: construire une liste et injecter le dernier message sans indexation négative sur QuerySet
    threads_list = list(threads)
    for t in threads_list:
        msgs = list(t.messages.all())
        t.last_message_obj = msgs[-1] if msgs else None

    return render(request, 'spot/correspondence_list.html', {
        'threads': threads,           # pour STATUS_CHOICES
        'threads_list': threads_list, # pour l'itération et l'aperçu
        'status': status,
        'q': q,
        'stats': stats,
    })

@login_required
def correspondence_thread(request, thread_id):
    """Détail d'une correspondance et ajout de messages"""
    thread = get_object_or_404(CorrespondenceThread, id=thread_id)

    # Contrôle d'accès
    if request.user.is_client() and thread.client != request.user and not request.user.is_admin():
        messages.error(request, "Accès non autorisé.")
        return redirect('correspondence_list')

    now = timezone.now()
    last_msg = thread.messages.last()
    can_reply_now = True
    next_allowed_reply_at = None

    # État dérivé (affichage)
    if thread.status == 'closed':
        state_code = 'closed'
        state_label = 'Résolu'
    elif last_msg and last_msg.author_id == thread.client_id:
        state_code = 'waiting'
        state_label = 'En attente de réponse'
    else:
        state_code = 'answered'
        state_label = 'Réponse reçue'

    # Délai 24h pour les relances client si pas de réponse du support
    if request.user.is_client():
        if last_msg and last_msg.author_id == request.user.id:
            limit = last_msg.created_at + timedelta(hours=24)
            if now < limit:
                can_reply_now = False
                next_allowed_reply_at = limit

    if request.method == 'POST':
        # Action: marquer comme résolu
        action = request.POST.get('action')
        if action == 'resolve':
            thread.status = 'closed'
            thread.save(update_fields=['status', 'updated_at'])
            messages.success(request, "Discussion marquée comme résolue.")
            return redirect('correspondence_thread', thread_id=thread.id)

        # Action: relancer (admin)
        if action == 'relance' and request.user.is_admin():
            relance_text = (request.POST.get('relance_text') or "Bonjour, nous relançons la discussion pour avancer.").strip()
            CorrespondenceMessage.objects.create(
                thread=thread,
                author=request.user,
                content=relance_text
            )
            thread.status = 'open'
            thread.last_message_at = timezone.now()
            thread.save(update_fields=['status', 'last_message_at', 'updated_at'])
            Notification.objects.create(
                user=thread.client,
                title='Relance du support',
                message=f'Le support a relancé la discussion: "{thread.subject}".',
                related_campaign=thread.related_campaign,
                related_thread=thread
            )
            messages.success(request, "Relance envoyée.")
            return redirect('correspondence_thread', thread_id=thread.id)

        # Envoi d'un message
        content = (request.POST.get('content') or '').strip()
        attachment = request.FILES.get('attachment')
        if not content:
            messages.error(request, "Veuillez saisir un message.")
        else:
            # Refuser la relance si délai non écoulé
            if request.user.is_client() and not can_reply_now:
                messages.warning(
                    request,
                    f"Vous pourrez relancer après le {next_allowed_reply_at.strftime('%d/%m/%Y %H:%M')}.")
                return redirect('correspondence_thread', thread_id=thread.id)

            CorrespondenceMessage.objects.create(
                thread=thread,
                author=request.user,
                content=content,
                attachment=attachment
            )
            # Mettre à jour statut + notifier
            if request.user.is_client():
                thread.status = 'pending'
                for admin in User.objects.filter(role='admin'):
                    Notification.objects.create(
                        user=admin,
                        title='Nouvelle réponse client',
                        message=f'{request.user.username} a répondu sur "{thread.subject}".',
                        related_campaign=thread.related_campaign,
                        related_thread=thread
                    )
            else:
                thread.status = 'open'
                Notification.objects.create(
                    user=thread.client,
                    title='Réponse du support',
                    message=f'Nouvelle réponse sur votre demande "{thread.subject}".',
                    related_campaign=thread.related_campaign,
                    related_thread=thread
                )
            thread.last_message_at = timezone.now()
            thread.save(update_fields=['status', 'last_message_at', 'updated_at'])

            messages.success(request, "Message envoyé.")
            return redirect('correspondence_thread', thread_id=thread.id)

    return render(request, 'spot/correspondence_thread.html', {
        'thread': thread,
        'messages_qs': thread.messages.all(),
        'can_reply_now': can_reply_now,
        'next_allowed_reply_at': next_allowed_reply_at,
        'state_code': state_code,
        'state_label': state_label,
    })

@login_required
def correspondence_new(request):
    """Création d'une nouvelle correspondance"""
    # Blocage préventif pour administrateurs
    if request.user.is_authenticated and hasattr(request.user, 'is_admin') and request.user.is_admin():
        logging.getLogger('bf1tv').info(
            'ADMIN_BLOCK_VIEW | user=%s username=%s view=correspondence_new method=%s at=%s',
            getattr(request.user, 'id', None), getattr(request.user, 'username', ''), request.method, timezone.now().isoformat(timespec='seconds')
        )
        messages.error(request, "Cette fonctionnalité est réservée aux clients. En tant qu'administrateur, vous n'avez pas accès à cette option.")
        return redirect('dashboard')
    if request.method == 'POST':
        subject = (request.POST.get('subject') or '').strip()
        content = (request.POST.get('content') or '').strip()
        campaign_id = request.POST.get('campaign_id')
        related_campaign = None
        if campaign_id:
            related_campaign = Campaign.objects.filter(id=campaign_id, client=request.user).first()

        if subject and content:
            thread = CorrespondenceThread.objects.create(
                client=request.user,
                subject=subject,
                related_campaign=related_campaign,
                status='open',
                last_message_at=timezone.now()
            )
            CorrespondenceMessage.objects.create(
                thread=thread,
                author=request.user,
                content=content
            )
            for admin in User.objects.filter(role='admin'):
                Notification.objects.create(
                    user=admin,
                    title='Nouvelle correspondance',
                    message=f'{request.user.username}: {subject}',
                    related_campaign=related_campaign,
                    related_thread=thread
                )
            messages.success(request, "Demande créée.")
            return redirect('correspondence_thread', thread_id=thread.id)
        else:
            messages.error(request, "Sujet et message sont requis.")

    campaigns = Campaign.objects.filter(client=request.user)
    return render(request, 'spot/correspondence_new.html', {
        'campaigns': campaigns
    })

@admin_required
def admin_correspondence_new(request):
    """Création d'une correspondance par un admin pour n'importe quel utilisateur/campagne."""
    if request.method == 'POST':
        subject = (request.POST.get('subject') or '').strip()
        content = (request.POST.get('content') or '').strip()
        client_id = request.POST.get('client_id')
        campaign_id = request.POST.get('campaign_id')

        client = User.objects.filter(id=client_id).first()
        campaign = Campaign.objects.filter(id=campaign_id).first() if campaign_id else None

        if not client:
            messages.error(request, "Veuillez choisir un utilisateur.")
        elif not subject or not content:
            messages.error(request, "Sujet et message sont requis.")
        else:
            thread = CorrespondenceThread.objects.create(
                client=client,
                subject=subject,
                related_campaign=campaign,
                status='open',
                last_message_at=timezone.now()
            )
            CorrespondenceMessage.objects.create(
                thread=thread,
                author=request.user,
                content=content
            )
            Notification.objects.create(
                user=client,
                title='Nouvelle discussion',
                message=f'Le support a démarré une discussion: "{subject}".',
                related_campaign=campaign,
                related_thread=thread
            )
            messages.success(request, "Discussion créée et utilisateur notifié.")
            return redirect('correspondence_thread', thread_id=thread.id)

    users = User.objects.all().order_by('username')
    campaigns = Campaign.objects.all().order_by('-created_at')
    return render(request, 'spot/correspondence_admin_new.html', {
        'users': users,
        'campaigns': campaigns,
    })

@login_required
def admin_login(request):
    """Connexion des administrateurs"""
    # Si déjà connecté et admin, aller au dashboard
    if request.user.is_authenticated and request.user.is_admin():
        return redirect('admin_dashboard')

    if request.method == 'POST':
        form = CustomAuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            if user and user.is_admin():
                login(request, user)
                messages.success(request, f'Bienvenue, {user.username} !')
                next_url = request.GET.get('next')
                return redirect(next_url) if next_url else redirect('admin_dashboard')
            else:
                messages.error(request, 'Accès refusé. Cette interface est réservée aux administrateurs.')
        else:
            messages.error(request, 'Identifiants invalides.')
    else:
        form = CustomAuthenticationForm()

    return render(request, 'spot/admin_login.html', {'form': form})

@login_required
def profile(request):
    """Page de profil utilisateur"""
    user = request.user
    
    if request.method == 'POST':
        # Traitement de la mise à jour du profil
        user.first_name = request.POST.get('first_name', user.first_name)
        user.last_name = request.POST.get('last_name', user.last_name)
        user.email = request.POST.get('email', user.email)
        user.phone = request.POST.get('phone', user.phone)
        user.company = request.POST.get('company', user.company)
        user.address = request.POST.get('address', user.address)
        
        # Changement de mot de passe (optionnel)
        password = request.POST.get('password')
        if password:
            user.set_password(password)
        
        user.save()
        messages.success(request, 'Votre profil a été mis à jour avec succès !')
        return redirect('profile')
    
    # Statistiques pour les clients
    stats = {}
    if user.is_client():
        stats = {
            'total_campaigns': Campaign.objects.filter(client=user).count(),
            'active_campaigns': Campaign.objects.filter(client=user, status='active').count(),
            'total_spots': Spot.objects.filter(campaign__client=user).count(),
            'pending_spots': Spot.objects.filter(campaign__client=user, status='pending_review').count(),
        }
    
    # Récupérer les dernières notifications
    notifications = Notification.objects.filter(user=user).order_by('-created_at')[:5]
    
    # Discussions et demandes initiées par l'utilisateur
    threads = CorrespondenceThread.objects.filter(client=user)\
        .select_related('related_campaign')\
        .order_by('-updated_at', '-created_at')[:10]
    contact_requests = ContactRequest.objects.filter(user=user)\
        .order_by('-updated_at', '-created_at')[:10]

    # Statistiques correspondance
    correspondence_stats = {
        'threads': {
            'open': CorrespondenceThread.objects.filter(client=user, status='open').count(),
            'pending': CorrespondenceThread.objects.filter(client=user, status='pending').count(),
            'closed': CorrespondenceThread.objects.filter(client=user, status='closed').count(),
        },
        'requests': {
            'new': ContactRequest.objects.filter(user=user, status='new').count(),
            'contacted': ContactRequest.objects.filter(user=user, status='contacted').count(),
            'closed': ContactRequest.objects.filter(user=user, status='closed').count(),
        },
    }
    
    context = {
        'user': user,
        'stats': stats,
        'notifications': notifications,
        'threads': threads,
        'contact_requests': contact_requests,
        'correspondence_stats': correspondence_stats,
    }
    
    return render(request, 'spot/profile.html', context)

# Module-level (ajouter sous les autres vues)
@login_required
def advisor_wizard(request):
    """Assistant d'orientation publicitaire: propose un format, une durée et un budget indicatif."""
    # Blocage préventif pour administrateurs
    if request.user.is_authenticated and hasattr(request.user, 'is_admin') and request.user.is_admin():
        logging.getLogger('bf1tv').info(
            'ADMIN_BLOCK_VIEW | user=%s username=%s view=advisor_wizard method=%s at=%s',
            getattr(request.user, 'id', None), getattr(request.user, 'username', ''), request.method, timezone.now().isoformat(timespec='seconds')
        )
        messages.error(request, "Cette fonctionnalité est réservée aux clients. En tant qu'administrateur, vous n'avez pas accès à cette option.")
        return redirect('dashboard')
    recommendation = None
    if request.method == 'POST':
        form = AdvisorWizardForm(request.POST)
        if form.is_valid():
            channel = form.cleaned_data['channel']
            has_spot_ready = form.cleaned_data['has_spot_ready'] == 'yes'
            objective = form.cleaned_data['objective']
            budget_estimate = form.cleaned_data.get('budget_estimate')

            duration_map = {'notoriete': 15, 'promotion': 20, 'sensibilisation': 30}
            recommended_duration = duration_map.get(objective, 15)
            recommended_medium = channel

            base_rule = PricingRule.objects.filter(is_active=True).order_by('base_price').first()
            if base_rule:
                duration_factor = Decimal(recommended_duration) / Decimal(15)
                time_factor = base_rule.time_slot_multiplier
                recommended_budget = (base_rule.base_price * duration_factor * time_factor).quantize(Decimal('1.'))
            else:
                recommended_budget = Decimal(budget_estimate or 0)

            AdvisorySession.objects.create(
                user=request.user if request.user.is_authenticated else None,
                channel=channel,
                has_spot_ready=has_spot_ready,
                objective=objective,
                budget_estimate=budget_estimate,
                recommended_medium=recommended_medium,
                recommended_duration_seconds=recommended_duration,
                recommended_budget=recommended_budget
            )

            recommendation = {
                'recommended_medium': recommended_medium,
                'recommended_duration': recommended_duration,
                'recommended_budget': recommended_budget,
                'has_spot_ready': has_spot_ready
            }
    else:
        form = AdvisorWizardForm()

    return render(request, 'spot/advisor_wizard.html', {'form': form, 'recommendation': recommendation})

@login_required



def contact_advisor(request):
    """Contact direct (Appel / WhatsApp) sans formulaire.
    Met l'accent sur l'action immédiate et propose des fallbacks.
    """
    # Blocage préventif pour administrateurs
    if request.user.is_authenticated and hasattr(request.user, 'is_admin') and request.user.is_admin():
        logging.getLogger('bf1tv').info(
            'ADMIN_BLOCK_VIEW | user=%s username=%s view=contact_advisor method=%s at=%s',
            getattr(request.user, 'id', None), getattr(request.user, 'username', ''), request.method, timezone.now().isoformat(timespec='seconds')
        )
        messages.error(request, "Cette fonctionnalité est réservée aux clients. En tant qu'administrateur, vous n'avez pas accès à cette option.")
        return redirect('dashboard')
    # Numéros de contact (paramétrables via settings, sinon valeurs par défaut)
    from django.conf import settings
    contact_phone = getattr(settings, 'BF1_CONTACT_PHONE', '+22674470032')
    whatsapp_phone = getattr(settings, 'BF1_WHATSAPP_PHONE', contact_phone)
    whatsapp_message = getattr(
        settings,
        'BF1_WHATSAPP_DEFAULT_MESSAGE',
        "Bonjour BF1TV, je souhaite démarrer une campagne publicitaire."
    )

    context = {
        'contact_phone': contact_phone,
        'whatsapp_phone': whatsapp_phone,
        'whatsapp_message': whatsapp_message,
    }
    return render(request, 'spot/contact_advisor.html', context)

def guides_list(request):
    """Liste des articles de conseil."""
    # Blocage préventif pour administrateurs
    if request.user.is_authenticated and hasattr(request.user, 'is_admin') and request.user.is_admin():
        logging.getLogger('bf1tv').info(
            'ADMIN_BLOCK_VIEW | user=%s username=%s view=guides_list method=%s at=%s',
            getattr(request.user, 'id', None), getattr(request.user, 'username', ''), request.method, timezone.now().isoformat(timespec='seconds')
        )
        messages.error(request, "Cette fonctionnalité est réservée aux clients. En tant qu'administrateur, vous n'avez pas accès à cette option.")
        return redirect('dashboard')
    articles = AdvisoryArticle.objects.filter(is_published=True).order_by('-published_at', '-created_at')
    curated_articles = [
        {
            'title': 'Choisir la bonne plage horaire',
            'slug': 'choisir-plage-horaire', 
            'summary': 'Repères simples: matin, midi, pré-access, prime time.',
            'tags': ['tv', 'jt'],
        },
        {
            'title': 'Formats et durées',
            'slug': 'formats-et-durees',
            'summary': 'Spots 0–30s vs 31–60s, suppléments, PAD.',
            'tags': ['tv', 'duree', 'format'],
        },
        {
            'title': 'Mixer TV et digital',
            'slug': 'mix-tv-digital',
            'summary': 'Associer diffusion TV et Facebook/YouTube.',
            'tags': ['digital', 'tv'],
        },
        {
            'title': 'Gérer un budget',
            'slug': 'gerer-le-budget',
            'summary': 'Combinaisons efficaces pour petit/moyen budget.',
            'tags': ['budget', 'tv', 'digital'],
        },
    ]
    return render(request, 'spot/guides_list.html', {
        'articles': articles,
        'curated_articles': curated_articles,
    })

def guide_detail(request, slug):
    """Détail d'un article."""
    # Blocage préventif pour administrateurs
    if request.user.is_authenticated and hasattr(request.user, 'is_admin') and request.user.is_admin():
        logging.getLogger('bf1tv').info(
            'ADMIN_BLOCK_VIEW | user=%s username=%s view=guide_detail slug=%s method=%s at=%s',
            getattr(request.user, 'id', None), getattr(request.user, 'username', ''), slug, request.method, timezone.now().isoformat(timespec='seconds')
        )
        messages.error(request, "Cette fonctionnalité est réservée aux clients. En tant qu'administrateur, vous n'avez pas accès à cette option.")
        return redirect('dashboard')
    article = AdvisoryArticle.objects.filter(slug=slug, is_published=True).first()
    return render(request, 'spot/guide_detail.html', {'article': article, 'slug': slug})

def inspiration(request):
    """Page d'inspiration: études de cas et témoignages."""
    # Blocage préventif pour administrateurs
    if request.user.is_authenticated and hasattr(request.user, 'is_admin') and request.user.is_admin():
        logging.getLogger('bf1tv').info(
            'ADMIN_BLOCK_VIEW | user=%s username=%s view=inspiration method=%s at=%s',
            getattr(request.user, 'id', None), getattr(request.user, 'username', ''), request.method, timezone.now().isoformat(timespec='seconds')
        )
        messages.error(request, "Cette fonctionnalité est réservée aux clients. En tant qu'administrateur, vous n'avez pas accès à cette option.")
        return redirect('dashboard')
    cases = CaseStudy.objects.filter(is_published=True).order_by('-published_at', '-created_at')
    return render(request, 'spot/inspiration.html', {'cases': cases})

@login_required
def report_overview(request):
    """Bilan intelligent: synthèse des campagnes sur une période (sans métriques Diffusions/Durée)."""
    start_date, end_date = _parse_period(request, default_to_current_month=True, max_span_days=365)

    # Campagnes filtrées par rôle et critères
    campaigns_qs = Campaign.objects.all() if request.user.is_admin() else Campaign.objects.filter(client=request.user)
    # Appliquer filtres multicritères si disponibles
    try:
        campaigns_qs = _apply_campaign_filters(campaigns_qs, request)
    except Exception:
        pass
    campaigns_month = campaigns_qs.filter(created_at__date__gte=start_date, created_at__date__lte=end_date).select_related('client')
    campaigns_count = campaigns_month.count()
    total_budget_month = campaigns_month.aggregate(total=Sum('budget'))['total'] or Decimal('0')

    # Demandes de couverture (période et rôle)
    coverages_qs = CoverageRequest.objects.select_related('user') if request.user.is_admin() else CoverageRequest.objects.filter(user=request.user).select_related('user')
    coverages_month = coverages_qs.filter(created_at__date__gte=start_date, created_at__date__lte=end_date)
    coverages_count = coverages_month.count()

    # Historique de consultation
    hist = request.session.get('report_history', [])
    hist.insert(0, {
        'start': start_date.isoformat(),
        'end': end_date.isoformat(),
        'q': request.GET.get('q') or '',
        'status': request.GET.get('status') or '',
        'channel': request.GET.get('channel') or '',
        'at': timezone.now().isoformat(timespec='seconds'),
    })
    request.session['report_history'] = hist[:10]

    return render(request, 'spot/report_overview.html', {
        'campaigns_count': campaigns_count,
        'start_date': start_date,
        'end_date': end_date,
        'campaigns_month': campaigns_month,
        'total_budget_month': total_budget_month,
        'export_history': request.session.get('export_history', [])[:5],
        'report_history': request.session.get('report_history', [])[:5],
        'coverages_month': coverages_month,
        'coverages_count': coverages_count,
    })

@login_required
def  excel_export_report(request):
    """Export Excel contenant exactement les informations du bilan (période + indicateurs + activités du mois)."""
    if Workbook is None:
        messages.error(request, "Veuillez installer openpyxl (pip install openpyxl) pour exporter en Excel.")
        return redirect('report_overview')

    # Période (identique à report_overview)
    start_date, end_date = _parse_period(request, default_to_current_month=True, max_span_days=365)

    # Campagnes filtrées par période et rôle
    campaigns_qs = Campaign.objects.all() if request.user.is_admin() else Campaign.objects.filter(client=request.user)
    try:
        campaigns_qs = _apply_campaign_filters(campaigns_qs, request)
    except Exception:
        pass
    campaigns_month = campaigns_qs.filter(created_at__date__gte=start_date, created_at__date__lte=end_date).select_related('client')
    campaigns_count = campaigns_month.count()
    total_budget_month = campaigns_month.aggregate(total=Sum('budget'))['total'] or Decimal('0')

    # Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Bilan"

    # Styles utilitaires (si disponibles)
    bold_font = Font(bold=True) if Font else None
    header_fill = PatternFill(start_color="FFE5E5", end_color="FFE5E5", fill_type="solid") if PatternFill else None
    center = Alignment(horizontal="center") if Alignment else None

    # En-tête du rapport et période
    ws["A1"] = "Bilan des activités"
    if bold_font:
        ws["A1"].font = bold_font

    ws["A2"] = "Période"
    ws["B2"] = f"{start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
    ws["A3"] = "Campagnes (total)"
    ws["B3"] = campaigns_count
    ws["A4"] = "Budget du mois"
    ws["B4"] = float(total_budget_month) if total_budget_month is not None else 0

    # Ligne de démarrage pour le tableau
    start_row = 8

    # En-tête du tableau activités (exactement les infos visibles)
    headers = ["Campagne", "Canal", "Client", "Statut", "Budget (FCFA)", "Date", "Lien"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=h)
        if bold_font:
            cell.font = bold_font
        if header_fill:
            cell.fill = header_fill
        if center:
            cell.alignment = center

    # Lignes du tableau
    row = start_row + 1
    for c in campaigns_month:
        # Campagne (titre)
        cell_title = ws.cell(row=row, column=1, value=c.title or "")
        if Alignment:
            cell_title.alignment = Alignment(wrap_text=True, shrink_to_fit=True)
        # Canal (label si disponible)
        chan = c.get_channel_display() if hasattr(c, 'get_channel_display') and c.channel else (c.channel or "")
        cell_chan = ws.cell(row=row, column=2, value=chan)
        if Alignment:
            cell_chan.alignment = Alignment(wrap_text=True, shrink_to_fit=True)
        # Client (société ou username)
        client_label = c.client.company if getattr(c.client, "company", None) else c.client.username
        cell_client = ws.cell(row=row, column=3, value=client_label)
        if Alignment:
            cell_client.alignment = Alignment(wrap_text=True, shrink_to_fit=True)
        # Statut (label)
        stat = c.get_status_display() if hasattr(c, 'get_status_display') else (getattr(c, 'status', '') or "")
        cell_status = ws.cell(row=row, column=4, value=stat)
        if Alignment:
            cell_status.alignment = Alignment(wrap_text=True, shrink_to_fit=True)
        # Budget
        ws.cell(row=row, column=5, value=float(c.budget or 0))
        # Date (création)
        ws.cell(row=row, column=6, value=c.created_at.strftime('%d/%m/%Y') if getattr(c, 'created_at', None) else "")
        # Lien "Voir" (hyperlien cliquable)
        campaign_url = request.build_absolute_uri(reverse('campaign_detail', args=[c.id]))
        link_cell = ws.cell(row=row, column=7, value="Voir")
        link_cell.hyperlink = campaign_url
        if Font:
            link_cell.font = Font(color="0000EE", underline="single")
        row += 1

    # Ajustement largeur de colonnes (pour lisibilité)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12

    # Validation: limiter longueur des colonnes textuelles (A-D)
    if DataValidation:
        dv_len = DataValidation(type="textLength", operator="lessThanOrEqual", formula1="120", allow_blank=True)
        ws.add_data_validation(dv_len)
        dv_len.add("A2:A1048576")
        dv_len.add("B2:B1048576")
        dv_len.add("C2:C1048576")
        dv_len.add("D2:D1048576")

    
    coverage_start_row = row + 2
    ws.cell(row=coverage_start_row, column=1, value="Demandes de couverture")
    if bold_font:
        ws.cell(row=coverage_start_row, column=1).font = bold_font

    cov_headers = ["Titre", "Type de couverture", "Date d’événement", "Contact", "Statut", "Date", "Lien"]
    for col_idx, h in enumerate(cov_headers, start=1):
        cell = ws.cell(row=coverage_start_row + 2, column=col_idx, value=h)
        if bold_font:
            cell.font = bold_font
        if header_fill:
            cell.fill = header_fill
        if center:
            cell.alignment = center

    cov_row = coverage_start_row + 3
    coverages_qs = CoverageRequest.objects.select_related('user') if request.user.is_admin() else CoverageRequest.objects.filter(user=request.user).select_related('user')
    coverages_month = coverages_qs.filter(created_at__date__gte=start_date, created_at__date__lte=end_date)
    for cv in coverages_month:
        ws.cell(row=cov_row, column=1, value=cv.event_title or "")
        ws.cell(row=cov_row, column=2, value=(cv.get_coverage_type_display() if hasattr(cv, 'get_coverage_type_display') and cv.coverage_type else (cv.coverage_type or "")))
        ws.cell(row=cov_row, column=3, value=cv.event_date.strftime('%d/%m/%Y') if getattr(cv, 'event_date', None) else "")
        ws.cell(row=cov_row, column=4, value=cv.contact_name or "")
        ws.cell(row=cov_row, column=5, value=getattr(cv, 'get_status_display')() if hasattr(cv, 'get_status_display') else (cv.status or ""))
        ws.cell(row=cov_row, column=6, value=cv.created_at.strftime('%d/%m/%Y') if getattr(cv, 'created_at', None) else "")
        coverage_url = request.build_absolute_uri(reverse('admin_coverage_detail', args=[cv.id])) if request.user.is_admin() else request.build_absolute_uri(reverse('coverage_detail', args=[cv.id]))
        link_cell = ws.cell(row=cov_row, column=7, value="Voir")
        link_cell.hyperlink = coverage_url
        if Font:
            link_cell.font = Font(color="0000EE", underline="single")
        cov_row += 1

    
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width, 30)
    ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width, 22)
    ws.column_dimensions["C"].width = max(ws.column_dimensions["C"].width, 14)
    ws.column_dimensions["D"].width = max(ws.column_dimensions["D"].width, 20)
    ws.column_dimensions["E"].width = max(ws.column_dimensions["E"].width, 15)
    ws.column_dimensions["F"].width = max(ws.column_dimensions["F"].width, 12)
    ws.column_dimensions["G"].width = max(ws.column_dimensions["G"].width, 12)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f"rapport_{start_date.isoformat()}_{end_date.isoformat()}.xlsx"
    response = HttpResponse(
        bio.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    # Historique des exports en session
    hist = request.session.get('export_history', [])
    hist.insert(0, {
        'format': 'excel',
        'filename': filename,
        'start': start_date.isoformat(),
        'end': end_date.isoformat(),
        'at': timezone.now().isoformat(timespec='seconds'),
    })
    request.session['export_history'] = hist[:20]
    return response

@login_required
def pdf_export_report(request):
    """Export PDF contenant les informations du bilan, avec une mise en page professionnelle."""
    # Vérifier la disponibilité de reportlab
    if SimpleDocTemplate is None:
        messages.error(request, "Veuillez installer reportlab (pip install reportlab) pour exporter en PDF.")
        return redirect('report_overview')

    # Période (identique à report_overview)
    start_date, end_date = _parse_period(request, default_to_current_month=True, max_span_days=365)

    campaigns_qs = Campaign.objects.all() if request.user.is_admin() else Campaign.objects.filter(client=request.user)
    try:
        campaigns_qs = _apply_campaign_filters(campaigns_qs, request)
    except Exception:
        pass
    campaigns_month = campaigns_qs.filter(created_at__date__gte=start_date, created_at__date__lte=end_date).select_related('client')
    campaigns_count = campaigns_month.count()
    total_budget_month = campaigns_month.aggregate(total=Sum('budget'))['total'] or Decimal('0')

    # Construction du PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, title="Rapport BF1 TV")
    styles = getSampleStyleSheet()
    story = []

    # Titre
    title = Paragraph("<b>Rapport des activités</b>", styles['Title'])
    story.append(title)
    story.append(Spacer(1, 12))

    # Période et indicateurs
    period = Paragraph(f"Période: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}", styles['Normal'])
    story.append(period)
    story.append(Spacer(1, 8))

    metrics_data = [
        ["Campagnes (total)", str(campaigns_count)],
        ["Budget du mois (FCFA)", f"{float(total_budget_month):,.0f}"],
    ]
    metrics_table = Table(metrics_data, colWidths=[200, 200])
    metrics_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#FFE5E5')),
        ('TEXTCOLOR', (0,0), (-1,-1), colors.black),
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,0), (-1,-1), 10),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('GRID', (0,0), (-1,-1), 0.25, colors.grey),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(metrics_table)
    story.append(Spacer(1, 16))

    # Tableau des activités du mois
    table_headers = ["Campagne", "Canal", "Client", "Statut", "Budget (FCFA)", "Date"]
    table_data = [table_headers]
    text_style = styles['Normal']
    # Forcer le wrap sur mots longs (URLs, mots collés)
    try:
        from reportlab.lib.styles import ParagraphStyle
        text_style = ParagraphStyle('Wrapped', parent=styles['Normal'], fontName='Helvetica', fontSize=9, leading=11)
    except Exception:
        pass

    truncated_any = False
    def trunc(txt, max_chars=200):
        nonlocal truncated_any
        if txt is None:
            return ""
        t = str(txt)
        if len(t) > max_chars:
            truncated_any = True
            return t[:max_chars-1] + '…'
        return t
    for c in campaigns_month:
        chan = c.get_channel_display() if hasattr(c, 'get_channel_display') and c.channel else (c.channel or "")
        client_label = c.client.company if getattr(c.client, "company", None) else c.client.username
        stat = c.get_status_display() if hasattr(c, 'get_status_display') else (getattr(c, 'status', '') or "")
        row = [
            Paragraph(trunc(c.title or ""), text_style),
            Paragraph(trunc(chan), text_style),
            Paragraph(trunc(client_label), text_style),
            Paragraph(trunc(stat), text_style),
            Paragraph(f"{float(c.budget or 0):,.0f}", text_style),
            Paragraph(c.created_at.strftime('%d/%m/%Y') if getattr(c, 'created_at', None) else "", text_style),
        ]
        table_data.append(row)

    activity_table = Table(table_data, repeatRows=1, colWidths=[110, 80, 100, 70, 80, 70])
    activity_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#F5F5F5')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.black),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('ALIGN', (0,0), (-1,0), 'CENTER'),
        ('ALIGN', (0,1), (-1,-1), 'LEFT'),
        ('GRID', (0,0), (-1,-1), 0.25, colors.grey),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#FAFAFA')]),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(activity_table)

    if truncated_any:
        story.append(Spacer(1, 8))
        story.append(Paragraph("Note: certains textes ont été tronqués (…).", styles['Italic']))

    
    story.append(Spacer(1, 16))
    story.append(Paragraph("<b>Demandes de couverture</b>", styles['Heading2']))

    coverages_qs = CoverageRequest.objects.select_related('user') if request.user.is_admin() else CoverageRequest.objects.filter(user=request.user).select_related('user')
    coverages_month = coverages_qs.filter(created_at__date__gte=start_date, created_at__date__lte=end_date)

    cov_headers = ["Titre", "Type de couverture", "Date d’événement", "Contact", "Statut", "Date"]
    cov_data = [cov_headers]
    for cv in coverages_month:
        stat = cv.get_status_display() if hasattr(cv, 'get_status_display') else (cv.status or "")
        cov_row = [
            Paragraph(trunc(cv.event_title or ""), text_style),
            Paragraph(trunc((cv.get_coverage_type_display() if hasattr(cv, 'get_coverage_type_display') and cv.coverage_type else (cv.coverage_type or ""))), text_style),
            Paragraph(cv.event_date.strftime('%d/%m/%Y') if getattr(cv, 'event_date', None) else "", text_style),
            Paragraph(trunc(cv.contact_name or ""), text_style),
            Paragraph(trunc(stat), text_style),
            Paragraph(cv.created_at.strftime('%d/%m/%Y') if getattr(cv, 'created_at', None) else "", text_style),
        ]
        cov_data.append(cov_row)

    if len(cov_data) > 1:
        cov_table = Table(cov_data, repeatRows=1, colWidths=[130, 100, 80, 110, 70, 60])
        cov_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#F5F5F5')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.black),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,0), (-1,-1), 9),
            ('ALIGN', (0,0), (-1,0), 'CENTER'),
            ('ALIGN', (0,1), (-1,-1), 'LEFT'),
            ('GRID', (0,0), (-1,-1), 0.25, colors.grey),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#FAFAFA')]),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ]))
        story.append(cov_table)
    else:
        story.append(Paragraph("Aucune demande de couverture pour la période sélectionnée.", styles['Normal']))

    # Générer PDF
    doc.build(story)
    buffer.seek(0)
    filename = f"rapport_{start_date.isoformat()}_{end_date.isoformat()}.pdf"
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Historique des exports en session
    hist = request.session.get('export_history', [])
    hist.insert(0, {
        'format': 'pdf',
        'filename': filename,
        'start': start_date.isoformat(),
        'end': end_date.isoformat(),
        'at': timezone.now().isoformat(timespec='seconds'),
    })
    request.session['export_history'] = hist[:20]
    return response

def pricing_overview(request):
    """Vue de la page de tarification publique"""
    # Blocage préventif pour administrateurs
    if request.user.is_authenticated and hasattr(request.user, 'is_admin') and request.user.is_admin():
        logging.getLogger('bf1tv').info(
            'ADMIN_BLOCK_VIEW | user=%s username=%s view=pricing_overview method=%s at=%s',
            getattr(request.user, 'id', None), getattr(request.user, 'username', ''), request.method, timezone.now().isoformat(timespec='seconds')
        )
        messages.error(request, "Cette fonctionnalité est réservée aux clients. En tant qu'administrateur, vous n'avez pas accès à cette option.")
        return redirect('dashboard')
    from .models import ServiceCategory, TimeSlot

    service_categories = ServiceCategory.objects.filter(is_active=True).prefetch_related('items').order_by('order', 'name')
    time_slots = TimeSlot.objects.filter(is_active=True).order_by('start_time')

    durations = [10, 15, 20, 30, 45, 60]  # affichage standard (indicatif)

    # Grille tarifaire extraite de l'image (montants illisibles marqués « Sur devis »)
    tariff_sheet = {
        "prestations": [
            {
                "no": 1,
                "label": "Confection et diffusion de petite annonce",
                "lines": [
                    {"sub": None, "unit": "JOUR", "number": 3, "price_ttc": 35400},
                    {"sub": None, "unit": "SEMAINE", "number": 21, "price_ttc": 200000},
                    {"sub": None, "unit": "MOIS", "number": 84, "price_ttc": 600000},
                ],
            },
            {
                "no": 2,
                "label": "Diffusion de message défilant",
                "lines": [
                    {"sub": "Diffusion", "unit": "JOUR", "number": 1, "price_ttc": 35400},
                    {"sub": "Diffusion", "unit": "SEMAINE", "number": 21, "price_ttc": 150000},
                ],
            },
            {
                "no": 3,
                "label": "Kosi Teedo",
                "lines": [
                    {"sub": "Réalisation et diffusion", "unit": "SEMAINE", "number": 17, "price_ttc": 300000},
                    {"sub": "Rediffusion", "unit": "SEMAINE", "number": 17, "price_ttc": None},  # Sur devis
                    {"sub": "Diffusion (≤ 6 min)", "unit": "UNITE", "number": 1, "price_ttc": 350000},
                ],
            },
            {
                "no": 4,
                "label": "Publireportage",
                "lines": [
                    {"sub": "Réalisation", "unit": "UNITE", "number": 1, "price_ttc": None},  # Sur devis
                    {"sub": "Diffusion JT 13h", "unit": "UNITE", "number": 1, "price_ttc": 250000},
                    {"sub": "Diffusion JT 19h30", "unit": "UNITE", "number": 1, "price_ttc": 350000},
                ],
            },
            {
                "no": 5,
                "label": "Passage aux émissions",
                "lines": [
                    {"sub": "La télé s'amuse - annonce", "unit": "UNITE", "number": 1, "price_ttc": 50000},
                    {"sub": "La télé s'amuse - avec public", "unit": "UNITE", "number": 1, "price_ttc": 150000},
                ],
            },
            {
                "no": 6,
                "label": "PAD (Diffusion)",
                "lines": [
                    {"sub": "≤ 13 min", "unit": "UNITE", "number": 1, "price_ttc": 500000},
                    {"sub": "≤ 26 min", "unit": "UNITE", "number": 1, "price_ttc": 750000},
                    {"sub": "≤ 52 min", "unit": "UNITE", "number": 1, "price_ttc": 1000000},
                ],
            },
            {
                "no": 7,
                "label": "Décrochage d'antenne (Réalisation)",
                "lines": [{"sub": None, "unit": "HEURE", "number": 1, "price_ttc": 1590000}],
            },
            {
                "no": 8,
                "label": "Réalisation d'émission spéciale en studio BF1",
                "lines": [{"sub": None, "unit": "HEURE", "number": 1, "price_ttc": 1130000}],
            },
            {
                "no": 9,
                "label": "Enregistrement d'émission spéciale hors studio à Ouaga",
                "lines": [{"sub": None, "unit": "HEURE", "number": 1, "price_ttc": 1500000}],
            },
            {
                "no": 10,
                "label": "Retransmission en direct hors studio à Ouaga",
                "lines": [{"sub": None, "unit": "HEURE", "number": 2, "price_ttc": 1100000}],  # à confirmer
            },
            {
                "no": 11,
                "label": "Retransmission en direct hors studio hors Ouaga",
                "lines": [{"sub": None, "unit": "HEURE", "number": 2, "price_ttc": None}],  # Sur devis
            },
            {
                "no": 12,
                "label": "Le clip à la Une",
                "lines": [
                    {"sub": "Sans message défilant", "unit": "SEMAINE", "number": 21, "price_ttc": 100000},
                    {"sub": "Avec message défilant", "unit": "SEMAINE", "number": 21, "price_ttc": 150000},
                    {"sub": "Avec 1 prestation à 'La télé s'amuse'", "unit": "SEMAINE", "number": 21, "price_ttc": None},  # Sur devis
                ],
            },
            {
                "no": 13,
                "label": "Diffusion sur la page Facebook/YouTube",
                "lines": [
                    {"sub": "Publireportage", "unit": "UNITE", "number": 1, "price_ttc": 150000},
                    {"sub": "Spot vidéo", "unit": "UNITE", "number": 1, "price_ttc": 50000},
                    {"sub": "Visuel", "unit": "UNITE", "number": 1, "price_ttc": 35400},
                ],
            },
        ],
        "spots": {
            "durations": ["0 à 30 s", "31 à 60 s"],
            "rows": [
                {"slot": "6h – 11h", "prices": [44870, 49855], "extra_per_second_ht": 997, "cultural_price": 35400},
                {"slot": "11h25 – 14h25", "prices": [103545, 115050], "extra_per_second_ht": 2301},
                {"slot": "14h30 – 16h55", "prices": [51773, 57525], "extra_per_second_ht": 1150},
                {"slot": "17h55 – 19h", "prices": [69030, 76700], "extra_per_second_ht": 1534},
                {"slot": "Prime time 19h25 – 20h55", "prices": [103545, 115050], "extra_per_second_ht": 2301, "prime": True},
                {"slot": "21h55 – fin", "prices": [69030, 76700], "extra_per_second_ht": 1534},
            ],
        },
        "coverage": {
            "title": "Couverture médiatique – activités ordinaires (Ouagadougou / Bobo-Dioulasso)",
            "rows": [
                {"label": "JT de 13h", "price": 250000},
                {"label": "JT de 19h30", "price": 350000},
                {"label": "Sport", "price": 150000},
            ],
        },
        "premium_offer": {
            "label": "JT 13h + JT 19h30 + Kibaye Wakato + rediffusion de chaque session",
            "price": 500000,
        },
    }

    context = {
        'service_categories': service_categories,
        'time_slots': time_slots,
        'durations': durations,
        'tariff_sheet': tariff_sheet
    }
    return render(request, 'spot/pricing_overview.html', context)

def ui_styleguide(request):
    """Documentation interne des composants UI (Tailwind)"""
    return render(request, 'spot/styleguide.html')
def _parse_period(request, default_to_current_month=True, max_span_days=365):
    """Parse `start`/`end` from query params with robust validation.
    - Fallback to current month if missing/invalid when `default_to_current_month` is True
    - Enforces start <= end and limits maximum span
    Returns (start_date, end_date)
    """
    start = request.GET.get('start')
    end = request.GET.get('end')
    try:
        start_date = datetime.strptime(start, '%Y-%m-%d').date() if start else None
        end_date = datetime.strptime(end, '%Y-%m-%d').date() if end else None
    except Exception:
        start_date, end_date = None, None

    if default_to_current_month and (not start_date or not end_date):
        today = timezone.localdate()
        month_start = today.replace(day=1)
        last_day = calendar.monthrange(today.year, today.month)[1]
        month_end = today.replace(day=last_day)
        start_date = start_date or month_start
        end_date = end_date or month_end

    if start_date and end_date and start_date > end_date:
        start_date, end_date = end_date, start_date

    if start_date and end_date and (end_date - start_date).days > max_span_days:
        end_date = start_date + timedelta(days=max_span_days)

    return start_date, end_date
def _apply_campaign_filters(qs, request):
    """Applique des filtres multicritères aux campagnes selon la requête.
    Filtre par texte (q), statut et canal si ces champs existent sur le modèle.
    """
    q = (request.GET.get('q') or '').strip()
    status = (request.GET.get('status') or '').strip()
    channel = (request.GET.get('channel') or '').strip()

    if q:
        qs = qs.filter(
            Q(title__icontains=q) |
            Q(client__username__icontains=q) |
            Q(client__company__icontains=q)
        )
    if status:
        try:
            qs = qs.filter(status=status)
        except Exception:
            pass
    if channel:
        try:
            qs = qs.filter(channel=channel)
        except Exception:
            pass
    return qs
@login_required
def editorial_dashboard(request):
    u = request.user
    if not (getattr(u, 'is_editorial_manager', lambda: False)() or getattr(u, 'is_admin', lambda: False)() or getattr(u, 'is_staff', False)):
        messages.error(request, 'Accès réservé à la rédaction.')
        return redirect('dashboard')
    today = timezone.now().date()
    pending = CoverageRequest.objects.filter(status='new').order_by('-created_at')[:10]
    todays = CoverageRequest.objects.filter(event_date=today).order_by('start_time')[:10]
    late = CoverageRequest.objects.filter(
        Q(response_deadline__lt=today) | Q(event_date__lt=today),
    ).exclude(status__in=['closed']).order_by('-event_date')[:10]
    journalists_available = Journalist.objects.filter(status='available').order_by('name')[:10]
    drivers_available = Driver.objects.filter(status='available').order_by('name')[:10]
    return render(request, 'editorial/dashboard.html', {
        'pending_coverages': pending,
        'todays_coverages': todays,
        'late_coverages': late,
        'journalists_available': journalists_available,
        'drivers_available': drivers_available,
        'today': today,
    })

@login_required
def editorial_coverage_detail(request, coverage_id):
    u = request.user
    if not (getattr(u, 'is_editorial_manager', lambda: False)() or getattr(u, 'is_admin', lambda: False)() or getattr(u, 'is_staff', False)):
        messages.error(request, 'Accès réservé à la rédaction.')
        return redirect('dashboard')
    coverage = get_object_or_404(CoverageRequest, id=coverage_id)
    assignments = coverage.assignments.select_related('journalist', 'driver').order_by('-assigned_at')
    # Simplifier: lister uniquement les disponibles pour une sélection claire
    spec = (coverage.event_type or '').lower()
    timeline = []
    timeline.append({'label': 'Demande reçue', 'at': coverage.created_at})
    if coverage.status in ['review', 'scheduled', 'closed']:
        timeline.append({'label': 'Validée', 'at': coverage.updated_at})
    for a in assignments:
        timeline.append({'label': 'Assignation', 'at': a.assigned_at})
        for lg in a.logs.all().order_by('at'):
            timeline.append({'label': lg.label, 'at': lg.at})
    all_journalists = Journalist.objects.all().order_by('name')
    all_drivers = Driver.objects.all().order_by('name')
    assigned_journalists = []
    assigned_drivers = []
    seen_j = set()
    seen_d = set()
    for a in assignments:
        if a.journalist and a.journalist_id and a.journalist_id not in seen_j:
            seen_j.add(a.journalist_id)
            assigned_journalists.append(a.journalist)
        if a.driver and a.driver_id and a.driver_id not in seen_d:
            seen_d.add(a.driver_id)
            assigned_drivers.append(a.driver)
    return render(request, 'editorial/coverage_detail.html', {
        'coverage': coverage,
        'assignments': assignments,
        'all_journalists': all_journalists,
        'all_drivers': all_drivers,
        'assigned_journalists': assigned_journalists,
        'assigned_drivers': assigned_drivers,
        'timeline': sorted(timeline, key=lambda x: x['at'] if x['at'] else timezone.now()),
    })

@login_required
def editorial_assign_coverage(request, coverage_id):
    u = request.user
    if not (getattr(u, 'is_editorial_manager', lambda: False)() or getattr(u, 'is_admin', lambda: False)() or getattr(u, 'is_staff', False)):
        messages.error(request, 'Accès réservé à la rédaction.')
        return redirect('dashboard')
    coverage = get_object_or_404(CoverageRequest, id=coverage_id)
    if coverage.status not in ['review', 'scheduled']:
        messages.error(request, "Cette couverture n'est pas encore validée par l'administration.")
        return redirect('editorial_coverage_detail', coverage_id=coverage.id)
    if request.method == 'POST':
        jids = request.POST.getlist('journalist_ids') or []
        dids = request.POST.getlist('driver_ids') or []
        if not jids and request.POST.get('journalist_id'):
            jids = [request.POST.get('journalist_id')]
        if not dids and request.POST.get('driver_id'):
            dids = [request.POST.get('driver_id')]

        created = []
        jids = [x for x in jids if x]
        dids = [x for x in dids if x]
        if len(jids) == 1 and len(dids) == 1:
            j = Journalist.objects.filter(id=jids[0]).first()
            d = Driver.objects.filter(id=dids[0]).first()
            if j and d:
                ass = CoverageAssignment.objects.create(coverage=coverage, journalist=j, driver=d, status='assigned')
                AssignmentLog.objects.create(assignment=ass, label='Assignation créée')
                j.status = 'on_mission'
                j.workload_score = (j.workload_score or 0) + 1
                j.save(update_fields=['status', 'workload_score', 'updated_at'])
                d.status = 'on_mission'
                d.save(update_fields=['status', 'updated_at'])
                created.append(ass)
        if not created:
            for jid in jids:
                j = Journalist.objects.filter(id=jid).first()
                if not j:
                    continue
                ass = CoverageAssignment.objects.create(coverage=coverage, journalist=j, driver=None, status='assigned')
                AssignmentLog.objects.create(assignment=ass, label='Assignation créée')
                j.status = 'on_mission'
                j.workload_score = (j.workload_score or 0) + 1
                j.save(update_fields=['status', 'workload_score', 'updated_at'])
                created.append(ass)

            for did in dids:
                d = Driver.objects.filter(id=did).first()
                if not d:
                    continue
                ass = CoverageAssignment.objects.create(coverage=coverage, journalist=None, driver=d, status='assigned')
                AssignmentLog.objects.create(assignment=ass, label='Assignation créée')
                d.status = 'on_mission'
                d.save(update_fields=['status', 'updated_at'])
                created.append(ass)

        if not created:
            messages.warning(request, "Aucun membre sélectionné.")
            return redirect('editorial_coverage_detail', coverage_id=coverage.id)

        admins = User.objects.filter(role='admin')
        for u in [u for u in admins] + ([request.user] if request.user else []):
            Notification.objects.create(user=u, title='Assignation couverture', message=f"{coverage.event_title}", type='success', related_coverage=coverage)

        try:
            from .utils import create_assignment_notification_campaigns
            for ass in created:
                create_assignment_notification_campaigns(ass, created_by=request.user)
        except Exception:
            logging.getLogger('spot').error('Erreur préparation notifications assignation', exc_info=True)

        messages.success(request, 'Assignation effectuée.')
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'ok': True, 'assignment_ids': [str(a.id) for a in created]})
    return redirect('editorial_coverage_detail', coverage_id=coverage.id)


def assignment_confirm(request, campaign_id, code):
    campaign = AssignmentNotificationCampaign.objects.select_related('assignment', 'assignment__coverage').filter(id=campaign_id).first()
    if not campaign or (campaign.confirm_code or '') != (code or ''):
        return HttpResponse('Code invalide.', status=404, content_type='text/plain')
    try:
        from .utils import confirm_assignment_notification_campaign
        confirm_assignment_notification_campaign(campaign, via='web')
    except Exception:
        logging.getLogger('spot').error('Erreur confirmation web assignation', exc_info=True)
    title = getattr(getattr(campaign.assignment, 'coverage', None), 'event_title', '')
    return HttpResponse(f'Confirmation enregistrée. {title}', content_type='text/plain')


def assignment_pdf(request, campaign_id, code):
    campaign = (
        AssignmentNotificationCampaign.objects.select_related('assignment', 'assignment__coverage')
        .filter(id=campaign_id)
        .first()
    )
    if not campaign or (campaign.confirm_code or '') != (code or ''):
        return HttpResponse('Code invalide.', status=404, content_type='text/plain')
    try:
        from .utils import build_coverage_pdf
        filename, content, mimetype = build_coverage_pdf(campaign.assignment.coverage)
    except Exception:
        logging.getLogger('spot').error('Erreur génération PDF assignation', exc_info=True)
        return HttpResponse('PDF indisponible.', status=500, content_type='text/plain')
    resp = HttpResponse(content, content_type=mimetype)
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


@login_required
@require_POST
def assignment_notify_email(request, campaign_id):
    u = request.user
    if not (getattr(u, 'is_editorial_manager', lambda: False)() or getattr(u, 'is_admin', lambda: False)() or getattr(u, 'is_staff', False)):
        messages.error(request, 'Accès réservé à la rédaction.')
        return redirect('dashboard')
    campaign = (
        AssignmentNotificationCampaign.objects.select_related('assignment', 'assignment__coverage')
        .filter(id=campaign_id)
        .first()
    )
    if not campaign:
        messages.error(request, "Notification introuvable.")
        return redirect('editorial_dashboard')
    if not campaign.to_email:
        messages.error(request, "Email indisponible pour ce destinataire.")
        return redirect('editorial_coverage_detail', coverage_id=campaign.assignment.coverage_id)
    try:
        from .models import AssignmentNotificationAttempt, AssignmentLog
        from .utils import build_coverage_pdf, send_assignment_notification_email
        pdf_att = build_coverage_pdf(campaign.assignment.coverage)
        recipient_label = ''
        if campaign.recipient_kind == 'journalist' and campaign.assignment.journalist:
            recipient_label = campaign.assignment.journalist.name
        if campaign.recipient_kind == 'driver' and campaign.assignment.driver:
            recipient_label = campaign.assignment.driver.name
        recipient_label = recipient_label or campaign.get_recipient_kind_display()
        ok = send_assignment_notification_email(
            to_email=campaign.to_email,
            recipient_label=recipient_label,
            subject='Assignation couverture',
            attachments=[pdf_att],
        )
        AssignmentNotificationAttempt.objects.create(
            campaign=campaign,
            channel='email',
            status='sent' if ok else 'failed',
            to=campaign.to_email,
            subject='Assignation couverture',
            body='PDF',
            sent_at=timezone.now() if ok else None,
            error='' if ok else 'envoi_failed',
        )
        if ok:
            campaign.status = 'confirmed'
            campaign.confirmed_at = timezone.now()
            campaign.confirmed_via = 'email'
            campaign.save(update_fields=['status', 'confirmed_at', 'confirmed_via', 'updated_at'])
            AssignmentLog.objects.create(assignment=campaign.assignment, label='Notification email envoyée', note=campaign.to_email)
            messages.success(request, 'Email envoyé.')
        else:
            messages.error(request, "Échec d'envoi email.")
    except Exception:
        logging.getLogger('spot').error('Erreur envoi email assignation', exc_info=True)
        messages.error(request, "Échec d'envoi email.")
    return redirect('editorial_coverage_detail', coverage_id=campaign.assignment.coverage_id)


@login_required
@require_POST
def assignment_notify_whatsapp(request, campaign_id):
    u = request.user
    if not (getattr(u, 'is_editorial_manager', lambda: False)() or getattr(u, 'is_admin', lambda: False)() or getattr(u, 'is_staff', False)):
        messages.error(request, 'Accès réservé à la rédaction.')
        return redirect('dashboard')
    campaign = (
        AssignmentNotificationCampaign.objects.select_related('assignment', 'assignment__coverage')
        .filter(id=campaign_id)
        .first()
    )
    if not campaign:
        messages.error(request, "Notification introuvable.")
        return redirect('editorial_dashboard')
    if not campaign.to_phone:
        messages.error(request, "Téléphone indisponible pour ce destinataire.")
        return redirect('editorial_coverage_detail', coverage_id=campaign.assignment.coverage_id)
    try:
        from .models import AssignmentNotificationAttempt, AssignmentLog
        pdf_url = request.build_absolute_uri(reverse('assignment_pdf', args=[campaign.id, campaign.confirm_code]))
        coverage = campaign.assignment.coverage
        title = (coverage.event_title or 'Couverture BF1').strip()
        text = f'Assignation BF1 TV — {title}\n\nFiche PDF complète:\n{pdf_url}'
        wa_link = f'https://wa.me/{campaign.to_phone.lstrip("+")}?text={urllib.parse.quote(text)}'
        AssignmentNotificationAttempt.objects.create(
            campaign=campaign,
            channel='whatsapp',
            status='sent',
            to=campaign.to_phone,
            body='PDF',
            sent_at=timezone.now(),
            meta={'link': wa_link, 'pdf_url': pdf_url},
        )
        campaign.status = 'confirmed'
        campaign.confirmed_at = timezone.now()
        campaign.confirmed_via = 'whatsapp'
        campaign.save(update_fields=['status', 'confirmed_at', 'confirmed_via', 'updated_at'])
        AssignmentLog.objects.create(assignment=campaign.assignment, label='Notification WhatsApp préparée', note=campaign.to_phone)
        return redirect(wa_link)
    except Exception:
        logging.getLogger('spot').error('Erreur notification WhatsApp assignation', exc_info=True)
        messages.error(request, "Échec préparation WhatsApp.")
        return redirect('editorial_coverage_detail', coverage_id=campaign.assignment.coverage_id)


@csrf_exempt
def sms_inbound(request):
    sender = (request.POST.get('from') or request.POST.get('From') or request.GET.get('from') or request.GET.get('From') or '').strip()
    body = (request.POST.get('body') or request.POST.get('Body') or request.POST.get('text') or request.POST.get('message') or request.GET.get('body') or request.GET.get('Body') or request.GET.get('text') or request.GET.get('message') or '').strip()
    m = re.search(r'\b(\d{6})\b', body or '')
    if not m:
        return JsonResponse({'ok': False, 'error': 'code_missing'}, status=400)
    code = m.group(1)
    qs = AssignmentNotificationCampaign.objects.filter(
        confirm_code=code,
        confirmed_at__isnull=True,
        status='active',
    ).order_by('-created_at')
    campaign = None
    if sender:
        try:
            from .utils import normalize_phone
            sn = normalize_phone(sender)
            tail = ''.join(ch for ch in sn if ch.isdigit())[-8:]
            if tail:
                for c in qs[:25]:
                    tp = ''.join(ch for ch in (c.to_phone or '') if ch.isdigit())[-8:]
                    if tp and tp == tail:
                        campaign = c
                        break
        except Exception:
            campaign = None
    if not campaign:
        campaign = qs.first()
    if not campaign:
        return JsonResponse({'ok': False, 'error': 'campaign_not_found'}, status=404)
    try:
        from .utils import confirm_assignment_notification_campaign
        confirm_assignment_notification_campaign(campaign, via='sms')
    except Exception:
        logging.getLogger('spot').error('Erreur confirmation SMS assignation', exc_info=True)
    return JsonResponse({'ok': True})

@login_required
def editorial_assignments(request):
    u = request.user
    if not (getattr(u, 'is_editorial_manager', lambda: False)() or getattr(u, 'is_admin', lambda: False)() or getattr(u, 'is_staff', False)):
        messages.error(request, 'Accès réservé à la rédaction.')
        return redirect('dashboard')
    if request.method == 'POST':
        action = request.POST.get('action')
        aid = request.POST.get('assignment_id')
        ass = CoverageAssignment.objects.select_related('coverage', 'journalist', 'driver').filter(id=aid).first()
        if ass:
            if action == 'close':
                ass.status = 'done'
                ass.save(update_fields=['status', 'updated_at'])
                AssignmentLog.objects.create(assignment=ass, label='Mission clôturée')
                if ass.journalist:
                    ass.journalist.status = 'available'
                    ass.journalist.save(update_fields=['status', 'updated_at'])
                if ass.driver:
                    ass.driver.status = 'available'
                    ass.driver.save(update_fields=['status', 'updated_at'])
                messages.success(request, 'Mission clôturée.')
            if action == 'upload':
                files = request.FILES.getlist('attachments')
                uploaded = 0
                for f in files:
                    try:
                        CoverageAttachment.objects.create(request=ass.coverage, file=f)
                        uploaded += 1
                    except Exception:
                        pass
                AssignmentLog.objects.create(assignment=ass, label='Contenu livré')
                if uploaded:
                    messages.success(request, f'{uploaded} fichier(s) uploadé(s).')
                else:
                    messages.warning(request, "Aucun fichier uploadé.")
        return redirect('editorial_assignments')
    q = (request.GET.get('q') or '').strip()
    mode = (request.GET.get('mode') or 'compact').strip()
    jid = (request.GET.get('journalist') or '').strip()
    did = (request.GET.get('driver') or '').strip()
    status = (request.GET.get('status') or '').strip()
    qs = CoverageAssignment.objects.select_related('coverage', 'journalist', 'driver')
    if q:
        qs = qs.filter(Q(coverage__event_title__icontains=q) | Q(coverage__address__icontains=q))
    if jid:
        qs = qs.filter(journalist__id=jid)
    if did:
        qs = qs.filter(driver__id=did)
    if status:
        qs = qs.filter(status=status)
    items = qs.order_by('-assigned_at')[:100]
    journalists = Journalist.objects.all().order_by('name')
    drivers = Driver.objects.all().order_by('name')
    return render(request, 'editorial/assignments.html', {
        'items': items,
        'journalists': journalists,
        'drivers': drivers,
        'q': q,
        'jid': jid,
        'did': did,
        'status': status,
        'mode': mode,
    })

@login_required
def editorial_notifications(request):
    u = request.user
    if not (getattr(u, 'is_editorial_manager', lambda: False)() or getattr(u, 'is_admin', lambda: False)() or getattr(u, 'is_staff', False)):
        messages.error(request, 'Accès réservé à la rédaction.')
        return redirect('dashboard')
    
    # Gestion des actions (marquer comme lu, archiver)
    if request.method == 'POST':
        action = request.POST.get('action')
        nid = request.POST.get('id')
        if action == 'mark_read':
            if nid:
                Notification.objects.filter(id=nid, user=request.user).update(is_read=True)
            else:
                Notification.objects.filter(user=request.user, is_read=False).update(is_read=True)
            return JsonResponse({'ok': True})
        elif action == 'delete':
            if nid:
                Notification.objects.filter(id=nid, user=request.user).delete()
            return JsonResponse({'ok': True})

    # Filtrer uniquement les notifications des couvertures validées par l'administration
    # On considère qu'une demande validée a le statut 'scheduled' ou 'review'
    # Et on filtre aussi par type 'success' (souvent utilisé pour les validations) ou 'info'
    qs = Notification.objects.filter(
        user=request.user,
        related_coverage__isnull=False,
        related_coverage__status__in=['scheduled', 'review']
    )
    
    # Filtres supplémentaires
    n_type = request.GET.get('type')
    date_filter = request.GET.get('date')
    
    if n_type:
        qs = qs.filter(type=n_type)
    if date_filter:
        try:
            d = datetime.strptime(date_filter, '%Y-%m-%d').date()
            qs = qs.filter(created_at__date=d)
        except Exception:
            pass
            
    items = qs.order_by('-created_at')[:100]
    return render(request, 'editorial/notifications.html', {'items': items})

@login_required
def editorial_planning(request):
    u = request.user
    if not (getattr(u, 'is_editorial_manager', lambda: False)() or getattr(u, 'is_admin', lambda: False)() or getattr(u, 'is_staff', False)):
        messages.error(request, 'Accès réservé à la rédaction.')
        return redirect('dashboard')
    view = (request.GET.get('view') or 'week').lower()
    ref_str = (request.GET.get('ref') or '').strip()
    try:
        base = datetime.strptime(ref_str, '%Y-%m-%d').date() if ref_str else timezone.localdate()
    except Exception:
        base = timezone.localdate()
    qs = CoverageRequest.objects.select_related('user').prefetch_related('assignments').exclude(status__in=['closed'])
    if view == 'day':
        target = base
        qs = qs.filter(event_date=target).order_by('start_time')
    elif view == 'month':
        start = base.replace(day=1)
        if start.month == 12:
            next_start = start.replace(year=start.year+1, month=1)
        else:
            next_start = start.replace(month=start.month+1)
        qs = qs.filter(event_date__gte=start, event_date__lt=next_start).order_by('event_date', 'start_time')
    else:
        start_week = base - timezone.timedelta(days=base.weekday())
        end_week = start_week + timezone.timedelta(days=7)
        qs = qs.filter(event_date__gte=start_week, event_date__lt=end_week).order_by('event_date', 'start_time')
    items = []
    for cv in qs:
        ass = cv.assignments.order_by('-assigned_at').first()
        items.append({
            'id': str(cv.id),
            'title': cv.event_title,
            'date': cv.event_date,
            'time': cv.start_time,
            'type': cv.event_type,
            'address': cv.address,
            'journalist': getattr(ass, 'journalist', None).name if ass and ass.journalist else '',
            'driver': getattr(ass, 'driver', None).name if ass and ass.driver else '',
            'status': cv.status,
            'detail_url': reverse('editorial_coverage_detail', args=[cv.id]),
        })
    week_days = [start_week + timezone.timedelta(days=i) for i in range(7)] if view == 'week' else []
    # Navigation dates
    if view == 'day':
        prev_ref = (base - timezone.timedelta(days=1)).strftime('%Y-%m-%d')
        next_ref = (base + timezone.timedelta(days=1)).strftime('%Y-%m-%d')
        range_label = base.strftime('%A %d/%m/%Y')
    elif view == 'month':
        prev_month = (base.replace(day=1) - timezone.timedelta(days=1)).replace(day=1)
        next_month_start = (base.replace(day=28) + timezone.timedelta(days=4)).replace(day=1)
        prev_ref = prev_month.strftime('%Y-%m-%d')
        next_ref = next_month_start.strftime('%Y-%m-%d')
        range_label = base.strftime('%B %Y')
    else:
        start_week = base - timezone.timedelta(days=base.weekday())
        end_week = start_week + timezone.timedelta(days=6)
        prev_ref = (start_week - timezone.timedelta(days=7)).strftime('%Y-%m-%d')
        next_ref = (start_week + timezone.timedelta(days=7)).strftime('%Y-%m-%d')
        range_label = f"Semaine du {start_week.strftime('%d/%m')} au {end_week.strftime('%d/%m')}"
    return render(request, 'editorial/planning.html', {
        'view': view,
        'items': items,
        'base': base,
        'week_days': week_days,
        'prev_ref': prev_ref,
        'next_ref': next_ref,
        'range_label': range_label,
    })

@login_required
def editorial_planning_move(request):
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'method_not_allowed'}, status=405)
    u = request.user
    if not (getattr(u, 'is_editorial_manager', lambda: False)() or getattr(u, 'is_admin', lambda: False)() or getattr(u, 'is_staff', False)):
        return JsonResponse({'ok': False, 'error': 'forbidden'}, status=403)
    cov_id = request.POST.get('coverage_id')
    new_date_str = request.POST.get('new_date')
    new_time_str = request.POST.get('new_time')
    try:
        coverage = CoverageRequest.objects.get(id=cov_id)
    except CoverageRequest.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'coverage_not_found'}, status=404)
    try:
        nd = datetime.strptime(new_date_str, '%Y-%m-%d').date() if new_date_str else coverage.event_date
    except Exception:
        nd = coverage.event_date
    try:
        nt = datetime.strptime(new_time_str, '%H:%M').time() if new_time_str else coverage.start_time
    except Exception:
        nt = coverage.start_time
    coverage.event_date = nd
    coverage.start_time = nt
    coverage.save(update_fields=['event_date', 'start_time', 'updated_at'])
    ass = coverage.assignments.order_by('-assigned_at').first()
    if ass:
        AssignmentLog.objects.create(assignment=ass, label='Planning déplacé')
    return JsonResponse({'ok': True})

@login_required
def editorial_coverages(request):
    u = request.user
    if not (getattr(u, 'is_editorial_manager', lambda: False)() or getattr(u, 'is_admin', lambda: False)() or getattr(u, 'is_staff', False)):
        messages.error(request, 'Accès réservé à la rédaction.')
        return redirect('dashboard')
    q = (request.GET.get('q') or '').strip()
    status = (request.GET.get('status') or '').strip() or 'validated'
    sort = (request.GET.get('sort') or 'date_desc').strip()
    page = request.GET.get('page')
    qs = CoverageRequest.objects.all().prefetch_related('assignments')
    if q:
        qs = qs.filter(Q(event_title__icontains=q) | Q(address__icontains=q) | Q(description__icontains=q))
    if status:
        if status == 'validated':
            qs = qs.filter(status__in=['review', 'scheduled'])
        else:
            qs = qs.filter(status=status)
    if sort == 'date_asc':
        qs = qs.order_by('event_date', 'start_time')
    else:
        qs = qs.order_by('-event_date', '-start_time')

    # N'afficher que les couvertures non assignées sur la page Couvertures
    qs = qs.annotate(assign_count=Count('assignments')).filter(assign_count=0)
    paginator = Paginator(qs, 24)
    page_obj = paginator.get_page(page)

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or (request.GET.get('format') == 'json'):
        data_items = []
        for cv in page_obj.object_list:
            data_items.append({
                'id': str(cv.id),
                'event_title': cv.event_title,
                'address': cv.address,
                'event_date': cv.event_date.strftime('%Y-%m-%d') if cv.event_date else None,
                'contact_name': cv.contact_name,
                'coverage_type': getattr(cv, 'get_coverage_type_display', lambda: cv.coverage_type)(),
                'detail_url': reverse('editorial_coverage_detail', args=[cv.id]),
            })
        return JsonResponse({
            'ok': True,
            'items': data_items,
            'pagination': {
                'has_next': page_obj.has_next(),
                'has_prev': page_obj.has_previous(),
                'num_pages': paginator.num_pages,
                'page': page_obj.number,
            }
        })

    status_counts = CoverageRequest.objects.values('status').annotate(count=Count('status'))
    counts = {item['status']: item['count'] for item in status_counts}
    inbox_items = CoverageRequest.objects.annotate(assign_count=Count('assignments')).filter(assign_count=0, status__in=['new','review']).order_by('event_date','start_time')[:12]
    available_journalists = Journalist.objects.all().order_by('name')
    available_drivers = Driver.objects.all().order_by('name')
    return render(request, 'editorial/coverages.html', {
        'page_obj': page_obj,
        'counts': counts,
        'q': q,
        'status': status,
        'sort': sort,
        'inbox_items': inbox_items,
        'available_journalists': available_journalists,
        'available_drivers': available_drivers,
    })
